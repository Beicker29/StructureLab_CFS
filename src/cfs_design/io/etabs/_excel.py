"""Small read-only Excel primitives restricted to the ETABS IO boundary."""

from collections.abc import Iterator
from contextlib import AbstractContextManager
from dataclasses import dataclass
from hashlib import sha256
from itertools import zip_longest
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from cfs_design.core.exceptions import ETABSImportError


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


@dataclass(frozen=True, slots=True)
class ExcelRow:
    source_path: Path
    worksheet: str
    row_number: int
    values: dict[str, object]
    formulas: dict[str, str]

    def error(self, field: str | None, message: str) -> ETABSImportError:
        field_text = f"; field {field!r}" if field is not None else ""
        return ETABSImportError(
            f"{self.source_path.name}; worksheet {self.worksheet!r}; "
            f"row {self.row_number}{field_text}: {message}"
        )


class ExcelWorkbookPair(AbstractContextManager["ExcelWorkbookPair"]):
    """Cached-value and formula views over one immutable workbook source."""

    def __init__(self, path: str | Path) -> None:
        supplied_path = Path(path).expanduser()
        if not supplied_path.is_file():
            raise ETABSImportError(f"Excel workbook does not exist: {supplied_path}")
        self.source_path = supplied_path.resolve()
        try:
            self.file_sha256 = sha256(self.source_path.read_bytes()).hexdigest()
        except OSError as error:
            raise ETABSImportError(
                f"Unable to read Excel workbook {self.source_path}: {error}"
            ) from error
        self.values: Workbook | None = None
        self.formulas: Workbook | None = None
        try:
            self.values = load_workbook(
                self.source_path, read_only=True, data_only=True
            )
            self.formulas = load_workbook(
                self.source_path, read_only=True, data_only=False
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
            if self.values is not None:
                self.values.close()
            if self.formulas is not None:
                self.formulas.close()
            raise ETABSImportError(
                f"Unable to read Excel workbook {self.source_path}: {error}"
            ) from error

    def __exit__(self, *_: object) -> None:
        if self.values is not None:
            self.values.close()
        if self.formulas is not None:
            self.formulas.close()

    def require_sheets(self, *sheet_names: str) -> None:
        assert self.formulas is not None
        missing = [name for name in sheet_names if name not in self.formulas.sheetnames]
        if missing:
            raise ETABSImportError(
                f"{self.source_path.name}: missing required worksheets: "
                f"{', '.join(missing)}"
            )

    def header_positions(
        self,
        worksheet: str,
        header_row: int,
        required_headers: tuple[str, ...],
    ) -> dict[str, int]:
        assert self.formulas is not None
        sheet = self.formulas[worksheet]
        raw_headers = next(
            sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ),
            (),
        )
        positions: dict[str, int] = {}
        duplicates: set[str] = set()
        for index, value in enumerate(raw_headers):
            if is_blank(value):
                continue
            if not isinstance(value, str):
                raise ETABSImportError(
                    f"{self.source_path.name}; worksheet {worksheet!r}; row "
                    f"{header_row}: column labels must be text"
                )
            label = value.strip()
            if label in positions:
                duplicates.add(label)
            positions[label] = index
        if duplicates:
            raise ETABSImportError(
                f"{self.source_path.name}; worksheet {worksheet!r}: duplicate "
                f"columns: {', '.join(sorted(duplicates))}"
            )
        missing = [header for header in required_headers if header not in positions]
        if missing:
            raise ETABSImportError(
                f"{self.source_path.name}; worksheet {worksheet!r}: missing "
                f"required columns: {', '.join(missing)}"
            )
        return positions

    def one_row(
        self,
        worksheet: str,
        row_number: int,
        positions: dict[str, int],
    ) -> ExcelRow:
        assert self.values is not None and self.formulas is not None
        cached = next(
            self.values[worksheet].iter_rows(
                min_row=row_number, max_row=row_number, values_only=True
            ),
            (),
        )
        formulas = next(
            self.formulas[worksheet].iter_rows(
                min_row=row_number, max_row=row_number, values_only=True
            ),
            (),
        )
        return self._make_row(worksheet, row_number, positions, cached, formulas)

    def rows(
        self,
        worksheet: str,
        start_row: int,
        positions: dict[str, int],
    ) -> Iterator[ExcelRow]:
        assert self.values is not None and self.formulas is not None
        cached_rows = self.values[worksheet].iter_rows(
            min_row=start_row, values_only=True
        )
        formula_rows = self.formulas[worksheet].iter_rows(
            min_row=start_row, values_only=True
        )
        for row_number, (cached, formulas) in enumerate(
            zip_longest(cached_rows, formula_rows, fillvalue=()),
            start=start_row,
        ):
            if all(is_blank(value) for value in formulas):
                continue
            yield self._make_row(
                worksheet, row_number, positions, cached, formulas
            )

    def _make_row(
        self,
        worksheet: str,
        row_number: int,
        positions: dict[str, int],
        cached: tuple[object, ...],
        formula_values: tuple[object, ...],
    ) -> ExcelRow:
        values = {
            header: cached[index] if index < len(cached) else None
            for header, index in positions.items()
        }
        formulas = {
            header: formula_values[index]
            for header, index in positions.items()
            if index < len(formula_values)
            and isinstance(formula_values[index], str)
            and formula_values[index].startswith("=")
        }
        return ExcelRow(
            source_path=self.source_path,
            worksheet=worksheet,
            row_number=row_number,
            values=values,
            formulas=formulas,
        )


def unresolved_formula(row: ExcelRow, field: str) -> None:
    if field in row.formulas and is_blank(row.values[field]):
        raise row.error(
            field,
            "formula cell has no usable cached value; formula evaluation is not supported",
        )


__all__ = [
    "ExcelRow",
    "ExcelWorkbookPair",
    "is_blank",
    "unresolved_formula",
]
