"""Validated loader for the full members.xlsx Members worksheet."""

from dataclasses import dataclass
from hashlib import sha256
from itertools import zip_longest
from math import isfinite
from numbers import Real
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from cfs_design.core.exceptions import ConfigurationError, SchemaError, ValidationError
from cfs_design.domain import (
    LengthDefinition,
    MemberCase,
    MemberGeometry,
    MemberType,
    Restraints,
)

from .models import MembersLoadResult, MembersWorkbookMetadata


LEGACY_MEMBERS_SCHEMA_VERSION = "0.1.0"
SUPPORTED_MEMBERS_SCHEMA_VERSION = "0.2.0"
SUPPORTED_MEMBERS_SCHEMA_VERSIONS = (
    LEGACY_MEMBERS_SCHEMA_VERSION,
    SUPPORTED_MEMBERS_SCHEMA_VERSION,
)
LEGACY_MEMBERS_COLUMNS = (
    "case_id",
    "label",
    "member_type",
    "section_id",
    "material_id",
    "L_mm",
    "length_definition",
    "Kx",
    "Ky",
    "Kt",
    "Lx_mm",
    "Ly_mm",
    "Lt_mm",
    "Lb_mm",
    "orientation_deg",
    "x_translation_restrained",
    "y_translation_restrained",
    "torsion_restrained",
    "warping_restrained",
    "lateral_brace_spacing_mm",
    "active",
    "notes",
)
MEMBERS_COLUMNS = LEGACY_MEMBERS_COLUMNS + (
    "distortional_unbraced_length_mm",
    "distortional_restraint_source",
)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


@dataclass(frozen=True, slots=True)
class _MemberRow:
    source_path: Path
    worksheet: str
    row_number: int
    values: dict[str, object]
    formulas: dict[str, str]

    def error(self, field: str | None, message: str) -> SchemaError:
        field_text = f"; field {field!r}" if field is not None else ""
        return SchemaError(
            f"{self.source_path.name}; worksheet {self.worksheet!r}; row "
            f"{self.row_number}{field_text}: {message}"
        )


class _MembersWorkbook:
    def __init__(self, path: str | Path) -> None:
        supplied = Path(path).expanduser()
        if not supplied.is_file():
            raise ConfigurationError(f"Members workbook does not exist: {supplied}")
        self.source_path = supplied.resolve()
        try:
            source_bytes = self.source_path.read_bytes()
            self.file_sha256 = sha256(source_bytes).hexdigest()
        except OSError as error:
            raise ConfigurationError(
                f"Unable to read members workbook {self.source_path}: {error}"
            ) from error
        self.values: Workbook | None = None
        self.formulas: Workbook | None = None
        try:
            self.values = load_workbook(
                self.source_path, read_only=True, data_only=True
            )
            self.formulas = load_workbook(
                self.source_path, read_only=True, data_only=False
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
            if self.values is not None:
                self.values.close()
            if self.formulas is not None:
                self.formulas.close()
            raise SchemaError(
                f"Unable to parse members workbook {self.source_path}: {error}"
            ) from error

    def __enter__(self) -> "_MembersWorkbook":
        return self

    def __exit__(self, *_: object) -> None:
        if self.values is not None:
            self.values.close()
        if self.formulas is not None:
            self.formulas.close()

    def require_sheets(self, *names: str) -> None:
        assert self.formulas is not None
        missing = [name for name in names if name not in self.formulas.sheetnames]
        if missing:
            raise SchemaError(
                f"{self.source_path.name}: missing required worksheets: "
                f"{', '.join(missing)}"
            )

    def header_positions(
        self,
        worksheet: str,
        headers: tuple[str, ...],
    ) -> dict[str, int]:
        assert self.formulas is not None
        raw = next(
            self.formulas[worksheet].iter_rows(
                min_row=1, max_row=1, values_only=True
            ),
            (),
        )
        positions: dict[str, int] = {}
        duplicates: set[str] = set()
        for index, value in enumerate(raw):
            if _is_blank(value):
                continue
            if not isinstance(value, str):
                raise SchemaError(
                    f"{self.source_path.name}; worksheet {worksheet!r}: "
                    "column labels must be text"
                )
            label = value.strip()
            if label in positions:
                duplicates.add(label)
            positions[label] = index
        if duplicates:
            raise SchemaError(
                f"{self.source_path.name}; worksheet {worksheet!r}: duplicate "
                f"columns: {', '.join(sorted(duplicates))}"
            )
        missing = [header for header in headers if header not in positions]
        if missing:
            raise SchemaError(
                f"{self.source_path.name}; worksheet {worksheet!r}: missing "
                f"required columns: {', '.join(missing)}"
            )
        return positions

    def rows(
        self,
        worksheet: str,
        positions: dict[str, int],
    ) -> tuple[_MemberRow, ...]:
        assert self.values is not None and self.formulas is not None
        cached_rows = self.values[worksheet].iter_rows(min_row=2, values_only=True)
        formula_rows = self.formulas[worksheet].iter_rows(
            min_row=2, values_only=True
        )
        rows: list[_MemberRow] = []
        for row_number, (cached, formulas) in enumerate(
            zip_longest(cached_rows, formula_rows, fillvalue=()), start=2
        ):
            if all(_is_blank(value) for value in formulas):
                continue
            values = {
                header: cached[index] if index < len(cached) else None
                for header, index in positions.items()
            }
            formula_cells = {
                header: formulas[index]
                for header, index in positions.items()
                if index < len(formulas)
                and isinstance(formulas[index], str)
                and formulas[index].startswith("=")
            }
            rows.append(
                _MemberRow(
                    source_path=self.source_path,
                    worksheet=worksheet,
                    row_number=row_number,
                    values=values,
                    formulas=formula_cells,
                )
            )
        return tuple(rows)


def _reject_formula(row: _MemberRow, field: str) -> None:
    if field in row.formulas:
        message = (
            "formula cell has no usable cached value; formula evaluation is not supported"
            if _is_blank(row.values[field])
            else "formula cells are not supported in the Members data contract"
        )
        raise row.error(
            field,
            message,
        )


def _required_text(row: _MemberRow, field: str) -> str:
    value = row.values[field]
    if not isinstance(value, str) or not value.strip():
        _reject_formula(row, field)
        raise row.error(field, f"invalid required text value: {value!r}")
    _reject_formula(row, field)
    return value


def _optional_text(row: _MemberRow, field: str) -> str | None:
    value = row.values.get(field)
    if _is_blank(value):
        _reject_formula(row, field)
        return None
    _reject_formula(row, field)
    if not isinstance(value, str):
        raise row.error(field, f"invalid optional text value: {value!r}")
    return value


def _required_number(row: _MemberRow, field: str) -> float:
    value = row.values[field]
    if isinstance(value, bool) or not isinstance(value, Real) or not isfinite(value):
        _reject_formula(row, field)
        raise row.error(field, f"invalid required numeric value: {value!r}")
    _reject_formula(row, field)
    return float(value)


def _optional_number(row: _MemberRow, field: str) -> float | None:
    value = row.values.get(field)
    if _is_blank(value):
        _reject_formula(row, field)
        return None
    return _required_number(row, field)


def _required_boolean(row: _MemberRow, field: str) -> bool:
    value = row.values[field]
    _reject_formula(row, field)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().upper()
        if normalized == "TRUE":
            return True
        if normalized == "FALSE":
            return False
    raise row.error(field, f"invalid boolean value: {value!r}")


def _required_enum(
    row: _MemberRow,
    field: str,
    enum_type: type[MemberType] | type[LengthDefinition],
) -> MemberType | LengthDefinition:
    value = _required_text(row, field)
    try:
        return enum_type(value)
    except ValueError as error:
        allowed = ", ".join(item.value for item in enum_type)
        raise row.error(
            field,
            f"unknown {enum_type.__name__} value {value!r}; expected: {allowed}",
        ) from error


def _metadata(reader: _MembersWorkbook) -> MembersWorkbookMetadata:
    positions = reader.header_positions("Metadata", ("Field", "Value"))
    values: dict[str, object] = {}
    order: list[str] = []
    for row in reader.rows("Metadata", positions):
        field = _required_text(row, "Field")
        if field in values:
            raise row.error("Field", f"duplicate metadata field {field!r}")
        value = row.values["Value"]
        _reject_formula(row, "Value")
        values[field] = value
        order.append(field)
    required_fields = ("name", "schema_version", "canonical_units")
    missing = [field for field in required_fields if field not in values]
    if missing:
        raise SchemaError(
            f"{reader.source_path.name}: Metadata is missing fields: "
            f"{', '.join(missing)}"
        )

    def text(field: str) -> str:
        value = values[field]
        if not isinstance(value, str) or not value.strip():
            raise SchemaError(
                f"{reader.source_path.name}: Metadata {field!r} must be text"
            )
        return value

    schema_version = text("schema_version")
    if schema_version not in SUPPORTED_MEMBERS_SCHEMA_VERSIONS:
        raise SchemaError(
            f"{reader.source_path.name}: unsupported schema_version "
            f"{schema_version!r}; supported versions are "
            f"{', '.join(repr(item) for item in SUPPORTED_MEMBERS_SCHEMA_VERSIONS)}"
        )
    return MembersWorkbookMetadata(
        name=text("name"),
        schema_version=schema_version,
        canonical_units=text("canonical_units"),
        source_path=reader.source_path,
        file_sha256=reader.file_sha256,
        additional_fields=tuple(
            (field, values[field]) for field in order if field not in required_fields
        ),
    )


def _member(row: _MemberRow) -> MemberCase:
    try:
        geometry = MemberGeometry(
            l_mm=_required_number(row, "L_mm"),
            length_definition=_required_enum(
                row, "length_definition", LengthDefinition
            ),  # type: ignore[arg-type]
            kx=_optional_number(row, "Kx"),
            ky=_optional_number(row, "Ky"),
            kt=_optional_number(row, "Kt"),
            lx_mm=_optional_number(row, "Lx_mm"),
            ly_mm=_optional_number(row, "Ly_mm"),
            lt_mm=_optional_number(row, "Lt_mm"),
            lb_mm=_optional_number(row, "Lb_mm"),
            orientation_deg=_required_number(row, "orientation_deg"),
        )
        restraints = Restraints(
            x_translation_restrained=_required_boolean(
                row, "x_translation_restrained"
            ),
            y_translation_restrained=_required_boolean(
                row, "y_translation_restrained"
            ),
            torsion_restrained=_required_boolean(row, "torsion_restrained"),
            warping_restrained=_required_boolean(row, "warping_restrained"),
            lateral_brace_spacing_mm=_optional_number(
                row, "lateral_brace_spacing_mm"
            ),
            distortional_unbraced_length_mm=_optional_number(
                row, "distortional_unbraced_length_mm"
            ),
            distortional_restraint_source=_optional_text(
                row, "distortional_restraint_source"
            ),
        )
        return MemberCase(
            case_id=_required_text(row, "case_id"),
            label=_required_text(row, "label"),
            member_type=_required_enum(
                row, "member_type", MemberType
            ),  # type: ignore[arg-type]
            section_id=_required_text(row, "section_id"),
            material_id=_required_text(row, "material_id"),
            geometry=geometry,
            restraints=restraints,
            active=_required_boolean(row, "active"),
            notes=_optional_text(row, "notes"),
        )
    except ValidationError as error:
        raise row.error(None, f"invalid MemberCase domain values: {error}") from error


def load_members(path: str | Path) -> MembersLoadResult:
    """Load all active and inactive physical member definitions."""

    with _MembersWorkbook(path) as reader:
        reader.require_sheets("Metadata", "Members")
        metadata = _metadata(reader)
        columns = (
            MEMBERS_COLUMNS
            if metadata.schema_version == SUPPORTED_MEMBERS_SCHEMA_VERSION
            else LEGACY_MEMBERS_COLUMNS
        )
        positions = reader.header_positions("Members", columns)
        members: list[MemberCase] = []
        seen: dict[str, int] = {}
        for row in reader.rows("Members", positions):
            member = _member(row)
            if member.case_id in seen:
                raise row.error(
                    "case_id",
                    f"duplicate case_id {member.case_id!r}; first seen at row "
                    f"{seen[member.case_id]}",
                )
            seen[member.case_id] = row.row_number
            members.append(member)
        if not members:
            raise SchemaError(
                f"{reader.source_path.name}: Members must contain at least one row"
            )
        return MembersLoadResult(metadata=metadata, members=tuple(members))


__all__ = [
    "LEGACY_MEMBERS_SCHEMA_VERSION",
    "MEMBERS_COLUMNS",
    "SUPPORTED_MEMBERS_SCHEMA_VERSION",
    "SUPPORTED_MEMBERS_SCHEMA_VERSIONS",
    "load_members",
]
