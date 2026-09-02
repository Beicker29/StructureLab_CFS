"""Shared, contextual Excel parsing restricted to the catalog boundary."""

from collections.abc import Iterator
from dataclasses import dataclass
from enum import Enum
from hashlib import sha256
from itertools import zip_longest
from math import isfinite
from numbers import Real
from pathlib import Path
from typing import TypeVar
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from cfs_design.core.exceptions import CatalogError, SchemaError

from .schemas import WorkbookSchema


EnumType = TypeVar("EnumType", bound=Enum)


@dataclass(frozen=True, slots=True)
class RowContext:
    source_path: Path
    worksheet: str
    row_number: int

    def catalog_error(self, field: str | None, message: str) -> CatalogError:
        field_line = f"\nField: {field}" if field is not None else ""
        return CatalogError(
            f"{self.source_path.name}\n"
            f"Worksheet: {self.worksheet}\n"
            f"Row: {self.row_number}"
            f"{field_line}\n\n"
            f"{message}"
        )


@dataclass(frozen=True, slots=True)
class CatalogRow:
    context: RowContext
    values: dict[str, object]
    formulas: dict[str, str]


class ExcelCatalogReader:
    """Validated read-only view over one catalog workbook."""

    def __init__(self, path: str | Path, schema: WorkbookSchema) -> None:
        supplied_path = Path(path).expanduser()
        if not supplied_path.is_file():
            raise CatalogError(f"Catalog workbook does not exist: {supplied_path}")
        self.source_path = supplied_path.resolve()
        try:
            self.file_sha256 = sha256(self.source_path.read_bytes()).hexdigest()
        except OSError as error:
            raise CatalogError(
                f"Unable to read catalog workbook {self.source_path}: {error}"
            ) from error
        self._value_workbook: Workbook | None = None
        self._formula_workbook: Workbook | None = None
        try:
            self._value_workbook = load_workbook(
                self.source_path,
                read_only=True,
                data_only=True,
            )
            self._formula_workbook = load_workbook(
                self.source_path,
                read_only=True,
                data_only=False,
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
            if self._value_workbook is not None:
                self._value_workbook.close()
            if self._formula_workbook is not None:
                self._formula_workbook.close()
            raise CatalogError(
                f"Unable to read catalog workbook {self.source_path}: {error}"
            ) from error
        self._column_positions: dict[str, dict[str, int]] = {}
        try:
            self._validate_schema(schema)
        except Exception:
            self._value_workbook.close()
            self._formula_workbook.close()
            raise

    def __enter__(self) -> "ExcelCatalogReader":
        return self

    def __exit__(self, *_: object) -> None:
        self._value_workbook.close()
        self._formula_workbook.close()

    def _validate_schema(self, schema: WorkbookSchema) -> None:
        missing_sheets = [
            sheet.name
            for sheet in schema.sheets
            if sheet.name not in self._formula_workbook.sheetnames
        ]
        if missing_sheets:
            raise SchemaError(
                f"{self.source_path.name}: missing required worksheets: "
                f"{', '.join(missing_sheets)}"
            )

        for sheet_schema in schema.sheets:
            worksheet = self._formula_workbook[sheet_schema.name]
            raw_headers = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            headers: list[str] = []
            positions: dict[str, int] = {}
            for index, value in enumerate(raw_headers):
                if value is None:
                    continue
                if not isinstance(value, str) or not value.strip():
                    raise SchemaError(
                        f"{self.source_path.name}: worksheet {sheet_schema.name} "
                        "contains an invalid column name"
                    )
                headers.append(value)
                positions[value] = index
            duplicates = sorted(
                {header for header in headers if headers.count(header) > 1}
            )
            if duplicates:
                raise SchemaError(
                    f"{self.source_path.name}: worksheet {sheet_schema.name} "
                    f"contains duplicate columns: {', '.join(duplicates)}"
                )
            missing_columns = [
                column for column in sheet_schema.columns if column not in headers
            ]
            if missing_columns:
                raise SchemaError(
                    f"{self.source_path.name}: worksheet {sheet_schema.name} "
                    f"is missing required columns: {', '.join(missing_columns)}"
                )
            self._column_positions[sheet_schema.name] = positions

    def validate_schema(self, schema: WorkbookSchema) -> None:
        """Validate an additional version-selected schema on the open workbook."""

        self._validate_schema(schema)

    def rows(self, worksheet_name: str) -> Iterator[CatalogRow]:
        value_worksheet = self._value_workbook[worksheet_name]
        formula_worksheet = self._formula_workbook[worksheet_name]
        positions = self._column_positions[worksheet_name]
        paired_rows = zip_longest(
            value_worksheet.iter_rows(min_row=2, values_only=True),
            formula_worksheet.iter_rows(min_row=2, values_only=True),
            fillvalue=(),
        )
        for row_number, (cached_values, formula_values) in enumerate(
            paired_rows,
            start=2,
        ):
            if all(is_blank(value) for value in formula_values):
                continue
            values = {
                header: cached_values[index] if index < len(cached_values) else None
                for header, index in positions.items()
            }
            formulas = {
                header: formula_values[index]
                for header, index in positions.items()
                if index < len(formula_values)
                and isinstance(formula_values[index], str)
                and formula_values[index].startswith("=")
            }
            yield CatalogRow(
                RowContext(self.source_path, worksheet_name, row_number),
                values,
                formulas,
            )


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def required_text(row: CatalogRow, field: str) -> str:
    value = row.values[field]
    if not isinstance(value, str) or not value.strip():
        _raise_unresolved_formula(row, field)
        raise row.context.catalog_error(field, f"Invalid required text value: {value!r}")
    return value


def optional_text(row: CatalogRow, field: str) -> str | None:
    value = row.values[field]
    if is_blank(value):
        _raise_unresolved_formula(row, field)
        return None
    if not isinstance(value, str):
        raise row.context.catalog_error(field, f"Invalid text value: {value!r}")
    return value


def required_number(row: CatalogRow, field: str) -> float:
    value = row.values[field]
    if (
        isinstance(value, bool)
        or not isinstance(value, Real)
        or not isfinite(value)
    ):
        _raise_unresolved_formula(row, field)
        raise row.context.catalog_error(field, f"Invalid numeric value: {value!r}")
    return float(value)


def required_integer(row: CatalogRow, field: str) -> int:
    value = required_number(row, field)
    if not value.is_integer():
        raise row.context.catalog_error(field, f"Invalid integer value: {value!r}")
    return int(value)


def optional_number(row: CatalogRow, field: str) -> float | None:
    value = row.values[field]
    if is_blank(value):
        return None
    return required_number(row, field)


def required_boolean(row: CatalogRow, field: str) -> bool:
    value = row.values[field]
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().upper()
        if normalized == "TRUE":
            return True
        if normalized == "FALSE":
            return False
    _raise_unresolved_formula(row, field)
    raise row.context.catalog_error(field, f"Invalid boolean value: {value!r}")


def _raise_unresolved_formula(row: CatalogRow, field: str) -> None:
    if field in row.formulas and is_blank(row.values[field]):
        raise row.context.catalog_error(
            field,
            "Formula cell has no usable cached value; formula evaluation is not supported",
        )


def required_enum(
    row: CatalogRow,
    field: str,
    enum_type: type[EnumType],
) -> EnumType:
    text = required_text(row, field)
    try:
        return enum_type(text)
    except ValueError as error:
        allowed = ", ".join(member.value for member in enum_type)
        raise row.context.catalog_error(
            field,
            f"Unknown {enum_type.__name__} value {text!r}; expected one of: {allowed}",
        ) from error


__all__ = [
    "CatalogRow",
    "ExcelCatalogReader",
    "RowContext",
    "optional_number",
    "optional_text",
    "required_boolean",
    "required_enum",
    "required_integer",
    "required_number",
    "required_text",
]
