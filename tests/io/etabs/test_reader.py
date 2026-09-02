"""Native ETABS reader tests against the approved workbook and invalid copies."""

from collections.abc import Callable
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cfs_design.core.exceptions import ETABSImportError
from cfs_design.io.etabs import ETABSImportConfig, read_etabs_results

APPROVED_ETABS_SHA256 = (
    "be2fc3b9b9d9fa57ca648fca533017bad6c4b572db2adbd7538d70b7cdd300ab"
)
REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
ETABS_SOURCE = REPOSITORY_ROOT / "projects" / "PRJ_001" / "ETABS_results.xlsx"


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _column(worksheet: object, header: str, row_number: int = 2) -> int:
    for cell in worksheet[row_number]:  # type: ignore[index]
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Test fixture has no column {header}")


def _set_force_cell(path: Path, row: int, header: str, value: object) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Element Forces - Beams"]  # type: ignore[index]
        worksheet.cell(row, _column(worksheet, header)).value = value

    _modify(path, operation)


def test_approved_workbook_metadata_units_and_all_raw_rows_are_preserved() -> None:
    result = read_etabs_results(ETABS_SOURCE)

    assert result.metadata.source_path == ETABS_SOURCE.resolve()
    assert result.metadata.file_sha256 == APPROVED_ETABS_SHA256
    assert result.metadata.file_sha256 == sha256(ETABS_SOURCE.read_bytes()).hexdigest()
    assert result.metadata.program_name == "ETABS"
    assert result.metadata.program_version == "18.1.1"
    assert result.metadata.current_units == "kip, in, F"
    assert not hasattr(result.metadata, "license_number")
    assert result.metadata.source_units.station == "m"
    assert result.metadata.source_units.p == "kgf"
    assert result.metadata.source_units.t == "kgf-m"
    assert len(result.raw_rows) == 24

    first = result.raw_rows[0]
    assert first.source_row == 4
    assert first.story == "P4B N+10.15"
    assert first.frame_label == "B114"
    assert first.unique_name == "1263"
    assert first.output_case == "DERX"
    assert first.case_type == "LinRespSpec"
    assert first.step_type == "Max"
    assert first.station_raw == 0.25
    assert first.p_raw == 0.0
    assert first.v2_raw == 12628.59
    assert first.t_raw == 279.05
    assert first.m3_raw == 31947.45


def test_config_is_explicit_and_immutable() -> None:
    config = ETABSImportConfig()
    assert config.header_row == 2
    assert config.units_row == 3
    assert config.data_start_row == 4
    assert config.columns.unique_name == "Unique Name"
    with pytest.raises(AttributeError):
        config.header_row = 3  # type: ignore[misc]


def test_blank_force_rows_are_ignored(etabs_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook["Element Forces - Beams"].insert_rows(5)  # type: ignore[index]

    _modify(etabs_copy, operation)
    result = read_etabs_results(etabs_copy)
    assert len(result.raw_rows) == 24
    assert result.raw_rows[1].source_row == 6


def test_partially_populated_force_row_is_rejected(etabs_copy: Path) -> None:
    _set_force_cell(etabs_copy, 4, "P", None)

    with pytest.raises(ETABSImportError, match="field 'P'.*invalid numeric"):
        read_etabs_results(etabs_copy)


def test_non_numeric_force_is_rejected_with_context(etabs_copy: Path) -> None:
    _set_force_cell(etabs_copy, 4, "V2", "not-a-number")

    with pytest.raises(ETABSImportError) as captured:
        read_etabs_results(etabs_copy)
    message = str(captured.value)
    assert "Element Forces - Beams" in message
    assert "row 4" in message
    assert "field 'V2'" in message


def test_formula_without_cached_value_is_rejected(etabs_copy: Path) -> None:
    _set_force_cell(etabs_copy, 4, "P", "=1+1")

    with pytest.raises(ETABSImportError, match="formula cell has no usable cached value"):
        read_etabs_results(etabs_copy)


def test_missing_required_force_column_is_rejected(etabs_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Element Forces - Beams"]  # type: ignore[index]
        worksheet.delete_cols(_column(worksheet, "M3"))

    _modify(etabs_copy, operation)
    with pytest.raises(ETABSImportError, match="M3"):
        read_etabs_results(etabs_copy)


def test_unknown_force_table_unit_is_rejected_even_if_program_units_exist(
    etabs_copy: Path,
) -> None:
    _set_force_cell(etabs_copy, 3, "P", "kN")

    with pytest.raises(ETABSImportError, match="Unsupported ETABS force-table units"):
        read_etabs_results(etabs_copy)


def test_missing_required_sheet_is_rejected(etabs_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook.remove(workbook["Program Control"])  # type: ignore[attr-defined,index]

    _modify(etabs_copy, operation)
    with pytest.raises(ETABSImportError, match="Program Control"):
        read_etabs_results(etabs_copy)
