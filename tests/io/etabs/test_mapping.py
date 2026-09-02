"""Exact, priority-based members.xlsx ETABS mapping tests."""

from collections.abc import Callable
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cfs_design.core.exceptions import ETABSImportError
from cfs_design.io.etabs import (
    import_etabs_results,
    load_etabs_mapping,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
ETABS_SOURCE = REPOSITORY_ROOT / "projects" / "PRJ_001" / "ETABS_results.xlsx"
MEMBERS_SOURCE = REPOSITORY_ROOT / "projects" / "PRJ_001" / "members.xlsx"
APPROVED_MEMBERS_SHA256 = (
    "288d2fe4bea7cbe514884db6fec52e5b5b13a433e2d3ac64a5d1136fad855ee8"
)


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _column(worksheet: object, header: str) -> int:
    for cell in worksheet[1]:  # type: ignore[index]
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Test fixture has no column {header}")


def _set_mapping(path: Path, row: int, **values: object) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["ETABS_Mapping"]  # type: ignore[index]
        for header, value in values.items():
            worksheet.cell(row, _column(worksheet, header)).value = value

    _modify(path, operation)


def test_approved_mapping_rows_are_preserved_but_disabled() -> None:
    mapping = load_etabs_mapping(MEMBERS_SOURCE)

    assert mapping.source_path == MEMBERS_SOURCE.resolve()
    assert mapping.file_sha256 == APPROVED_MEMBERS_SHA256
    assert mapping.file_sha256 == sha256(MEMBERS_SOURCE.read_bytes()).hexdigest()
    assert len(mapping.rows) == 2
    assert mapping.enabled_rows == ()
    assert mapping.rows[0].case_id == "EX_BEAM_001"
    assert mapping.rows[0].etabs_unique_name == "1263"
    assert mapping.rows[0].notes is not None


def test_unique_name_has_first_mapping_priority(members_copy: Path) -> None:
    _set_mapping(members_copy, 2, mapping_enabled=True)
    mapping = load_etabs_mapping(members_copy)
    result = import_etabs_results(ETABS_SOURCE, mapping=mapping)

    assert tuple(member.case_id for member in result.mapped_members) == (
        "EX_BEAM_001",
    )
    assert len(result.mapped_members[0].records) == 24
    assert result.unmapped_rows == ()


def test_story_and_beam_are_exact_fallback_when_unique_name_is_blank(
    members_copy: Path,
) -> None:
    _set_mapping(
        members_copy,
        2,
        etabs_unique_name=None,
        mapping_enabled=True,
    )
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )

    assert len(result.mapped_members) == 1
    assert len(result.mapped_members[0].records) == 24


def test_controlled_surrounding_whitespace_is_ignored(members_copy: Path) -> None:
    _set_mapping(
        members_copy,
        2,
        etabs_unique_name="  1263  ",
        etabs_story="DOES NOT MATCH",
        etabs_beam="DOES NOT MATCH",
        mapping_enabled=True,
    )
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )
    assert len(result.mapped_members[0].records) == 24


def test_disabled_rows_are_ignored_and_demands_remain_visible() -> None:
    mapping = load_etabs_mapping(MEMBERS_SOURCE)
    result = import_etabs_results(ETABS_SOURCE, mapping=mapping)

    assert result.mapped_members == ()
    assert len(result.unmapped_rows) == 24
    assert result.warnings == ("Unmapped ETABS demand rows: 24",)


def test_unknown_etabs_object_is_not_silently_discarded(members_copy: Path) -> None:
    _set_mapping(
        members_copy,
        2,
        etabs_unique_name="UNKNOWN",
        etabs_story="UNKNOWN STORY",
        etabs_beam="UNKNOWN BEAM",
        mapping_enabled=True,
    )
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )

    assert result.mapped_members == ()
    assert len(result.unmapped_rows) == 24
    assert any("matched no ETABS rows" in warning for warning in result.warnings)
    assert any("Unmapped ETABS demand rows: 24" == warning for warning in result.warnings)


def test_conflicting_unique_and_fallback_matches_are_rejected(
    members_copy: Path,
) -> None:
    _set_mapping(
        members_copy,
        2,
        etabs_story="OTHER STORY",
        etabs_beam="OTHER BEAM",
        mapping_enabled=True,
    )
    _set_mapping(
        members_copy,
        3,
        etabs_unique_name="OTHER UNIQUE",
        etabs_story="P4B N+10.15",
        etabs_beam="B114",
        mapping_enabled=True,
    )
    mapping = load_etabs_mapping(members_copy)

    with pytest.raises(ETABSImportError, match="Ambiguous ETABS mapping"):
        import_etabs_results(ETABS_SOURCE, mapping=mapping)


def test_duplicate_enabled_unique_name_is_rejected(members_copy: Path) -> None:
    _set_mapping(members_copy, 2, mapping_enabled=True)
    _set_mapping(
        members_copy,
        3,
        etabs_unique_name="1263",
        mapping_enabled=True,
    )

    with pytest.raises(ETABSImportError, match="Unique Name.*maps more than once"):
        load_etabs_mapping(members_copy)


def test_duplicate_enabled_story_and_beam_is_rejected(members_copy: Path) -> None:
    _set_mapping(members_copy, 2, mapping_enabled=True)
    _set_mapping(
        members_copy,
        3,
        etabs_story="P4B N+10.15",
        etabs_beam="B114",
        mapping_enabled=True,
    )

    with pytest.raises(ETABSImportError, match=r"Story \+ Beam.*maps more than once"):
        load_etabs_mapping(members_copy)


def test_one_case_id_cannot_map_to_two_etabs_objects(members_copy: Path) -> None:
    _set_mapping(members_copy, 2, mapping_enabled=True)
    _set_mapping(
        members_copy,
        3,
        case_id="EX_BEAM_001",
        mapping_enabled=True,
    )

    with pytest.raises(ETABSImportError, match="case_id.*duplicated"):
        load_etabs_mapping(members_copy)


def test_one_mapping_cannot_aggregate_distinct_etabs_unique_names(
    members_copy: Path,
    etabs_copy: Path,
) -> None:
    _set_mapping(members_copy, 2, mapping_enabled=True)

    def change_one_unique_name(workbook: object) -> None:
        worksheet = workbook["Element Forces - Beams"]  # type: ignore[index]
        headers = {cell.value: cell.column for cell in worksheet[2]}
        worksheet.cell(4, headers["Unique Name"]).value = "OTHER FRAME"

    _modify(etabs_copy, change_one_unique_name)
    mapping = load_etabs_mapping(members_copy)

    with pytest.raises(ETABSImportError, match="one-to-many aggregation"):
        import_etabs_results(etabs_copy, mapping=mapping)


def test_enabled_mapping_requires_a_complete_strategy(members_copy: Path) -> None:
    _set_mapping(
        members_copy,
        2,
        etabs_unique_name=None,
        etabs_beam=None,
        mapping_enabled=True,
    )

    with pytest.raises(ETABSImportError, match="supplied together"):
        load_etabs_mapping(members_copy)
