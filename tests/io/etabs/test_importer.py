"""End-to-end M4 import regression without member/design resolution."""

from collections.abc import Callable
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cfs_design.io.etabs import import_etabs_results, load_etabs_mapping


REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
ETABS_SOURCE = REPOSITORY_ROOT / "projects" / "PRJ_001" / "ETABS_results.xlsx"


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _enable_first_mapping(path: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["ETABS_Mapping"]  # type: ignore[index]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        worksheet.cell(2, headers["mapping_enabled"], True)

    _modify(path, operation)


def test_import_preserves_output_cases_stations_locations_and_every_row(
    members_copy: Path,
) -> None:
    _enable_first_mapping(members_copy)
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )

    assert len(result.raw_rows) == 24
    assert len(result.normalized_rows) == 24
    assert len(result.mapped_members) == 1
    member = result.mapped_members[0]
    assert len(member.records) == 24
    assert tuple(
        combination.combination_id for combination in member.demand_set.combinations
    ) == ("DERX", "DERY")
    assert tuple(
        combination.case_type for combination in member.demand_set.combinations
    ) == ("LinRespSpec", "LinRespSpec")
    assert tuple(
        len(combination.points) for combination in member.demand_set.combinations
    ) == (12, 12)

    for combination in member.demand_set.combinations:
        assert {point.step_type for point in combination.points} == {"Max"}
        discontinuity = tuple(
            point
            for point in combination.points
            if point.station_mm == pytest.approx(2750.0)
        )
        assert len(discontinuity) == 2
        assert {point.location for point in discontinuity} == {"Before", "After"}
        assert all(point.element_station_mm == point.station_mm for point in combination.points)

    assert result.unmapped_rows == ()
    assert result.warnings == ()


def test_response_spectrum_rows_are_not_enveloped_or_sign_fabricated(
    members_copy: Path,
) -> None:
    _enable_first_mapping(members_copy)
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )
    member = result.mapped_members[0]

    assert len(member.records) == len(result.raw_rows) == 24
    assert tuple(record.raw_row.source_row for record in member.records) == tuple(
        range(4, 28)
    )
    assert all(record.raw_row.case_type == "LinRespSpec" for record in member.records)
    assert all(record.raw_row.step_type == "Max" for record in member.records)
    assert not any(record.raw_row.step_type == "Min" for record in member.records)


def test_each_normalized_point_has_exact_workbook_and_row_provenance(
    members_copy: Path,
) -> None:
    _enable_first_mapping(members_copy)
    result = import_etabs_results(
        ETABS_SOURCE,
        mapping=load_etabs_mapping(members_copy),
    )

    assert result.metadata.source_path == ETABS_SOURCE.resolve()
    assert result.metadata.file_sha256.startswith("be2fc3b9b9d9")
    for record in result.normalized_rows:
        assert record.raw_row.worksheet == result.metadata.force_sheet
        assert record.raw_row.source_row >= result.metadata.data_start_row
        assert record.demand_point.point_id.endswith(
            f"R{record.raw_row.source_row:06d}"
        )


def test_import_without_mapping_keeps_all_rows_explicitly_unmapped() -> None:
    result = import_etabs_results(ETABS_SOURCE)

    assert result.mapping is None
    assert result.mapped_members == ()
    assert len(result.unmapped_rows) == 24
    assert "No ETABS mapping supplied" in result.warnings[0]
