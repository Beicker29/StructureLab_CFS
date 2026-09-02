"""M5 full Members worksheet loader tests."""

from collections.abc import Callable
from hashlib import sha256
from pathlib import Path
from shutil import copy2

import pytest
from openpyxl import load_workbook

from cfs_design.core.exceptions import SchemaError
from cfs_design.domain import LengthDefinition, MemberType
from cfs_design.io.project import load_members


REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
MEMBERS_SOURCE = REPOSITORY_ROOT / "projects" / "PRJ_001" / "members.xlsx"
APPROVED_MEMBERS_SHA256 = (
    "5b6add4838f33d0ad49f286f16ede5655bc9a681e4ba29fb4192ad182485f6f2"
)


@pytest.fixture
def members_copy(tmp_path: Path) -> Path:
    target = tmp_path / "members.xlsx"
    copy2(MEMBERS_SOURCE, target)
    return target


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _column(worksheet: object, header: str) -> int:
    for cell in worksheet[1]:  # type: ignore[index]
        if cell.value == header:
            return cell.column
    raise AssertionError(f"No column {header}")


def _set(path: Path, row: int, header: str, value: object) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Members"]  # type: ignore[index]
        worksheet.cell(row, _column(worksheet, header)).value = value

    _modify(path, operation)


def _upgrade_to_schema_02(path: Path) -> None:
    def operation(workbook: object) -> None:
        metadata = workbook["Metadata"]  # type: ignore[index]
        for row in range(2, metadata.max_row + 1):
            if metadata.cell(row, 1).value == "schema_version":
                metadata.cell(row, 2).value = "0.2.0"
                break
        else:
            raise AssertionError("No schema_version metadata row")
        members = workbook["Members"]  # type: ignore[index]
        headers = {cell.value for cell in members[1]}
        if "distortional_unbraced_length_mm" not in headers:
            members.cell(1, members.max_column + 1).value = (
                "distortional_unbraced_length_mm"
            )
        headers = {cell.value for cell in members[1]}
        if "distortional_restraint_source" not in headers:
            members.cell(1, members.max_column + 1).value = (
                "distortional_restraint_source"
            )

    _modify(path, operation)


def test_approved_members_loads_both_inactive_member_styles() -> None:
    result = load_members(MEMBERS_SOURCE)

    assert result.metadata.schema_version == "0.2.0"
    assert result.metadata.source_path == MEMBERS_SOURCE.resolve()
    assert result.metadata.file_sha256 == APPROVED_MEMBERS_SHA256
    assert result.metadata.file_sha256 == sha256(MEMBERS_SOURCE.read_bytes()).hexdigest()
    assert len(result.members) == 2
    assert result.active_members == ()
    first, second = result.members
    assert first.case_id == "EX_BEAM_001"
    assert first.member_type is MemberType.BEAM
    assert first.geometry.length_definition is LengthDefinition.K_FACTORS
    assert first.geometry.kx == first.geometry.ky == first.geometry.kt == 1.0
    assert second.geometry.length_definition is LengthDefinition.EFFECTIVE_LENGTHS
    assert second.geometry.lx_mm == 4000.0
    assert first.restraints.y_translation_restrained is True
    assert first.restraints.distortional_unbraced_length_mm is None
    assert first.restraints.distortional_restraint_source is None


def test_duplicate_case_id_is_rejected_with_rows(members_copy: Path) -> None:
    _set(members_copy, 3, "case_id", "EX_BEAM_001")
    with pytest.raises(SchemaError, match="duplicate case_id.*first seen at row 2"):
        load_members(members_copy)


@pytest.mark.parametrize("value", ("yes", 1, 0, None))
def test_invalid_boolean_is_rejected(members_copy: Path, value: object) -> None:
    _set(members_copy, 2, "active", value)
    with pytest.raises(SchemaError, match="field 'active'.*invalid boolean"):
        load_members(members_copy)


@pytest.mark.parametrize(
    ("field", "value", "match"),
    (
        ("member_type", "FRAME", "unknown MemberType"),
        ("length_definition", "AUTO", "unknown LengthDefinition"),
    ),
)
def test_invalid_member_enum_is_rejected(
    members_copy: Path,
    field: str,
    value: str,
    match: str,
) -> None:
    _set(members_copy, 2, field, value)
    with pytest.raises(SchemaError, match=match):
        load_members(members_copy)


def test_malformed_partial_row_is_rejected(members_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Members"]  # type: ignore[index]
        values: list[object | None] = [None] * worksheet.max_column
        values[_column(worksheet, "case_id") - 1] = "PARTIAL"
        worksheet.append(values)

    _modify(members_copy, operation)
    with pytest.raises(SchemaError, match="invalid required"):
        load_members(members_copy)


def test_complete_blank_rows_are_ignored(members_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook["Members"].insert_rows(3)  # type: ignore[index]

    _modify(members_copy, operation)
    assert len(load_members(members_copy).members) == 2


def test_optional_blanks_remain_none(members_copy: Path) -> None:
    result = load_members(members_copy)
    assert result.members[0].geometry.lx_mm is None
    assert result.members[1].geometry.kx is None


def test_m1_intrinsic_member_validation_is_reused(members_copy: Path) -> None:
    _set(members_copy, 2, "L_mm", -1.0)
    with pytest.raises(SchemaError, match="invalid MemberCase domain values.*greater than zero"):
        load_members(members_copy)


def test_contradictory_length_definition_is_rejected(members_copy: Path) -> None:
    _set(members_copy, 2, "Lx_mm", 5500.0)
    with pytest.raises(SchemaError, match="contradictory effective lengths"):
        load_members(members_copy)


def test_formula_without_cached_value_is_rejected(members_copy: Path) -> None:
    _set(members_copy, 2, "L_mm", "=5000+500")
    with pytest.raises(SchemaError, match="formula cell has no usable cached value"):
        load_members(members_copy)


def test_missing_required_column_is_rejected(members_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Members"]  # type: ignore[index]
        worksheet.delete_cols(_column(worksheet, "orientation_deg"))

    _modify(members_copy, operation)
    with pytest.raises(SchemaError, match="orientation_deg"):
        load_members(members_copy)


def test_schema_02_preserves_valid_explicit_lm_and_source(
    members_copy: Path,
) -> None:
    _upgrade_to_schema_02(members_copy)
    _set(members_copy, 2, "distortional_unbraced_length_mm", 1800.0)
    _set(
        members_copy,
        2,
        "distortional_restraint_source",
        "Structural restraint schedule R-01.",
    )

    result = load_members(members_copy)
    restraints = result.members[0].restraints

    assert result.metadata.schema_version == "0.2.0"
    assert restraints.distortional_unbraced_length_mm == 1800.0
    assert restraints.distortional_restraint_source == (
        "Structural restraint schedule R-01."
    )


def test_schema_02_allows_missing_lm_when_not_required(members_copy: Path) -> None:
    _upgrade_to_schema_02(members_copy)

    restraints = load_members(members_copy).members[0].restraints

    assert restraints.distortional_unbraced_length_mm is None
    assert restraints.distortional_restraint_source is None


def test_legacy_schema_01_loads_without_fabricating_lm(
    members_copy: Path,
) -> None:
    def operation(workbook: object) -> None:
        metadata = workbook["Metadata"]  # type: ignore[index]
        for row in range(2, metadata.max_row + 1):
            if metadata.cell(row, 1).value == "schema_version":
                metadata.cell(row, 2).value = "0.1.0"
                break
        members = workbook["Members"]  # type: ignore[index]
        for header in reversed(
            (
                "distortional_unbraced_length_mm",
                "distortional_restraint_source",
            )
        ):
            members.delete_cols(_column(members, header))

    _modify(members_copy, operation)
    result = load_members(members_copy)

    assert result.metadata.schema_version == "0.1.0"
    assert all(
        member.restraints.distortional_unbraced_length_mm is None
        and member.restraints.distortional_restraint_source is None
        for member in result.members
    )


@pytest.mark.parametrize("value", (0.0, -1.0, "Infinity", "NaN"))
def test_schema_02_rejects_invalid_lm(
    members_copy: Path,
    value: object,
) -> None:
    _upgrade_to_schema_02(members_copy)
    _set(members_copy, 2, "distortional_unbraced_length_mm", value)
    _set(
        members_copy,
        2,
        "distortional_restraint_source",
        "Synthetic test source.",
    )

    with pytest.raises(SchemaError, match="distortional_unbraced_length_mm"):
        load_members(members_copy)


def test_lm_is_never_inferred_from_existing_bracing_fields(
    members_copy: Path,
) -> None:
    _upgrade_to_schema_02(members_copy)
    _set(members_copy, 2, "Lb_mm", 1600.0)
    _set(members_copy, 2, "lateral_brace_spacing_mm", 900.0)

    restraints = load_members(members_copy).members[0].restraints

    assert restraints.distortional_unbraced_length_mm is None
    assert restraints.distortional_restraint_source is None


def test_schema_02_rejects_unpaired_lm_source(members_copy: Path) -> None:
    _upgrade_to_schema_02(members_copy)
    _set(members_copy, 2, "distortional_unbraced_length_mm", 1800.0)

    with pytest.raises(SchemaError, match="must be supplied together"):
        load_members(members_copy)
