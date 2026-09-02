"""M8A.2 standard-specific material qualification contract tests."""

from collections.abc import Callable
from dataclasses import FrozenInstanceError
from pathlib import Path
from shutil import copy2

import pytest
from openpyxl import load_workbook

from cfs_design.catalogs import load_material_catalog, load_section_catalog
from cfs_design.catalogs.registry import CatalogRegistry
from cfs_design.catalogs.schemas import MATERIAL_QUALIFICATION_COLUMNS
from cfs_design.core.exceptions import CatalogError, SchemaError
from cfs_design.domain import (
    A3ElongationGroup,
    MaterialQualificationRoute,
    MaterialQualificationState,
    StandardMaterialQualification,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
MATERIALS_SOURCE = REPOSITORY_ROOT / "data" / "catalogs" / "materials_catalog.xlsx"
SECTIONS_SOURCE = REPOSITORY_ROOT / "data" / "catalogs" / "sections_catalog.xlsx"


@pytest.fixture
def qualification_catalog(tmp_path: Path) -> Path:
    target = tmp_path / "materials_catalog.xlsx"
    copy2(MATERIALS_SOURCE, target)
    _modify(target, _upgrade_to_schema_02)
    return target


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _headers(worksheet: object) -> dict[object, int]:
    return {cell.value: cell.column for cell in worksheet[1]}  # type: ignore[index]


def _upgrade_to_schema_02(workbook: object) -> None:
    metadata = workbook["Metadata"]  # type: ignore[index]
    headers = _headers(metadata)
    for row in range(2, metadata.max_row + 1):
        if metadata.cell(row, headers["Field"]).value == "schema_version":
            metadata.cell(row, headers["Value"]).value = "0.2.0"
            break
    else:
        raise AssertionError("Metadata has no schema_version row")
    if "AISI_Material_Qualification" not in workbook.sheetnames:  # type: ignore[attr-defined]
        worksheet = workbook.create_sheet(  # type: ignore[attr-defined]
            "AISI_Material_Qualification", 2
        )
        worksheet.append(MATERIAL_QUALIFICATION_COLUMNS)
    else:
        worksheet = workbook["AISI_Material_Qualification"]  # type: ignore[index]
        assert tuple(cell.value for cell in worksheet[1]) == (
            MATERIAL_QUALIFICATION_COLUMNS
        )


def _downgrade_to_schema_01(workbook: object) -> None:
    workbook.remove(workbook["AISI_Material_Qualification"])  # type: ignore[attr-defined,index]
    metadata = workbook["Metadata"]  # type: ignore[index]
    headers = _headers(metadata)
    for row in range(2, metadata.max_row + 1):
        if metadata.cell(row, headers["Field"]).value == "schema_version":
            metadata.cell(row, headers["Value"]).value = "0.1.0"
            return
    raise AssertionError("Metadata has no schema_version row")


def _valid_values(**replacements: object) -> dict[str, object]:
    values: dict[str, object] = {
        "material_id": "EX_MAT_G50",
        "standard_id": "ANSI_SDI_AISI_S100",
        "standard_edition": 2024,
        "qualification_route": "A3_1",
        "qualification_state": "QUALIFIED",
        "product_form": "SHEET",
        "steel_classification": "CARBON",
        "elongation_group": "A3_1_1_GE_10",
        "minimum_elongation_percent": 10.0,
        "elongation_gauge_length_mm": 50.0,
        "elongation_test_standard": "ASTM_A370",
        "mandatory_mechanical_properties_state": "SATISFIED",
        "test_reports_required_state": "SATISFIED",
        "chemical_mechanical_conformance_state": "NOT_APPLICABLE",
        "properties_determined_per_reference_state": "NOT_APPLICABLE",
        "coating_requirements_state": "NOT_APPLICABLE",
        "welding_requirements_state": "NOT_APPLICABLE",
        "production_identification_state": "NOT_APPLICABLE",
        "master_coil_10_percent_overstrength_state": "NOT_APPLICABLE",
        "local_elongation_percent": None,
        "uniform_elongation_percent": None,
        "ductility_test_standard": None,
        "source_id": "EX_SRC_MAT",
        "basis": "Synthetic test evidence; not production qualification.",
        "notes": "SYNTHETIC_TEST_DATA",
    }
    values.update(replacements)
    return values


def _append_record(path: Path, **replacements: object) -> None:
    values = _valid_values(**replacements)

    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Material_Qualification"]  # type: ignore[index]
        worksheet.append([values[column] for column in MATERIAL_QUALIFICATION_COLUMNS])

    _modify(path, operation)


def test_valid_qualification_is_typed_immutable_and_lookupable(
    qualification_catalog: Path,
) -> None:
    _append_record(qualification_catalog)

    catalog = load_material_catalog(qualification_catalog)
    qualification = catalog.get_material_qualification(
        "EX_MAT_G50", "ANSI_SDI_AISI_S100", 2024
    )

    assert isinstance(qualification, StandardMaterialQualification)
    assert qualification.qualification_route is MaterialQualificationRoute.A3_1
    assert qualification.qualification_state is MaterialQualificationState.QUALIFIED
    assert qualification.elongation_group is A3ElongationGroup.A3_1_1_GE_10
    with pytest.raises(FrozenInstanceError):
        qualification.basis = "changed"  # type: ignore[misc]


def test_legacy_catalog_loads_with_explicitly_missing_qualification(
    tmp_path: Path,
) -> None:
    legacy = tmp_path / "materials_catalog_legacy.xlsx"
    copy2(MATERIALS_SOURCE, legacy)
    _modify(legacy, _downgrade_to_schema_01)

    catalog = load_material_catalog(legacy)

    assert catalog.metadata.schema_version == "0.1.0"
    assert catalog.material_qualifications == ()
    assert catalog.find_material_qualification(
        "EX_MAT_G50", "ANSI_SDI_AISI_S100", 2024
    ) is None
    with pytest.raises(CatalogError, match="Unknown material qualification"):
        catalog.get_material_qualification(
            "EX_MAT_G50", "ANSI_SDI_AISI_S100", 2024
        )


def test_registry_exposes_exact_qualification_lookup(
    qualification_catalog: Path,
) -> None:
    _append_record(qualification_catalog)
    registry = CatalogRegistry(
        load_material_catalog(qualification_catalog),
        load_section_catalog(SECTIONS_SOURCE),
    )

    assert registry.find_material_qualification(
        "EX_MAT_G50", "ANSI_SDI_AISI_S100", 2024
    ) is registry.get_material_qualification(
        "EX_MAT_G50", "ANSI_SDI_AISI_S100", 2024
    )


def test_schema_02_requires_qualification_worksheet(
    qualification_catalog: Path,
) -> None:
    _modify(
        qualification_catalog,
        lambda workbook: workbook.remove(  # type: ignore[attr-defined]
            workbook["AISI_Material_Qualification"]  # type: ignore[index]
        ),
    )
    with pytest.raises(SchemaError, match="AISI_Material_Qualification"):
        load_material_catalog(qualification_catalog)


def test_schema_02_requires_all_qualification_columns(
    qualification_catalog: Path,
) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Material_Qualification"]  # type: ignore[index]
        worksheet.delete_cols(_headers(worksheet)["qualification_route"])

    _modify(qualification_catalog, operation)
    with pytest.raises(SchemaError, match="qualification_route"):
        load_material_catalog(qualification_catalog)


def test_duplicate_composite_key_is_rejected(qualification_catalog: Path) -> None:
    _append_record(qualification_catalog)
    _append_record(qualification_catalog)
    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_material_catalog(qualification_catalog)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    (
        ("material_id", "UNKNOWN", "Materials.material_id"),
        ("source_id", "UNKNOWN", "Sources.source_id"),
        ("standard_edition", 2016, "Unsupported material-qualification"),
        ("standard_id", "OTHER", "Unsupported material-qualification"),
        ("qualification_route", "A3_FAKE", "MaterialQualificationRoute"),
    ),
)
def test_invalid_reference_standard_or_route_is_rejected(
    qualification_catalog: Path,
    field: str,
    value: object,
    message: str,
) -> None:
    _append_record(qualification_catalog, **{field: value})
    with pytest.raises(CatalogError, match=message):
        load_material_catalog(qualification_catalog)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    (
        ("minimum_elongation_percent", None, "require elongation"),
        ("elongation_gauge_length_mm", 100.0, "A3.1.1 requires"),
        ("test_reports_required_state", "UNKNOWN", "test_reports_required_state"),
    ),
)
def test_qualified_a3_1_requires_complete_route_evidence(
    qualification_catalog: Path,
    field: str,
    value: object,
    message: str,
) -> None:
    _append_record(qualification_catalog, **{field: value})
    with pytest.raises(CatalogError, match=message):
        load_material_catalog(qualification_catalog)


def test_valid_a3_2_record_accepts_documented_production_route(
    qualification_catalog: Path,
) -> None:
    _append_record(
        qualification_catalog,
        qualification_route="A3_2",
        mandatory_mechanical_properties_state="NOT_APPLICABLE",
        test_reports_required_state="NOT_APPLICABLE",
        chemical_mechanical_conformance_state="SATISFIED",
        properties_determined_per_reference_state="SATISFIED",
        coating_requirements_state="NOT_APPLICABLE",
        welding_requirements_state="NOT_APPLICABLE",
        production_identification_state="SATISFIED",
    )

    qualification = load_material_catalog(
        qualification_catalog
    ).material_qualifications[0]
    assert qualification.qualification_route is MaterialQualificationRoute.A3_2


def test_a3_2_missing_conditional_evidence_is_rejected(
    qualification_catalog: Path,
) -> None:
    _append_record(
        qualification_catalog,
        qualification_route="A3_2",
        mandatory_mechanical_properties_state="NOT_APPLICABLE",
        test_reports_required_state="NOT_APPLICABLE",
        chemical_mechanical_conformance_state="UNKNOWN",
    )
    with pytest.raises(CatalogError, match="chemical_mechanical_conformance_state"):
        load_material_catalog(qualification_catalog)


def test_partial_row_is_rejected(qualification_catalog: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Material_Qualification"]  # type: ignore[index]
        worksheet.append(["EX_MAT_G50"])

    _modify(qualification_catalog, operation)
    with pytest.raises(CatalogError, match="Invalid required text value"):
        load_material_catalog(qualification_catalog)


def test_blank_row_is_ignored(qualification_catalog: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Material_Qualification"]  # type: ignore[index]
        worksheet.append([None] * len(MATERIAL_QUALIFICATION_COLUMNS))

    _modify(qualification_catalog, operation)
    assert load_material_catalog(qualification_catalog).material_qualifications == ()


def test_any_formula_is_rejected(qualification_catalog: Path) -> None:
    _append_record(qualification_catalog, minimum_elongation_percent="=5+5")
    with pytest.raises(CatalogError, match="Formulas are prohibited"):
        load_material_catalog(qualification_catalog)


def test_a3_1_1_uses_existing_material_strengths_for_ratio(
    qualification_catalog: Path,
) -> None:
    _append_record(qualification_catalog)

    def operation(workbook: object) -> None:
        worksheet = workbook["Materials"]  # type: ignore[index]
        worksheet.cell(3, _headers(worksheet)["Fu_MPa"]).value = 350.0

    _modify(qualification_catalog, operation)
    with pytest.raises(CatalogError, match="Fu/Fy >= 1.08"):
        load_material_catalog(qualification_catalog)
