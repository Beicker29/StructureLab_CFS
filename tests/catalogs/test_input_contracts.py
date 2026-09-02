"""Read-only smoke tests for approved workbook contracts.

These tests deliberately validate structure, not engineering content or a
complete loader. Workbook access must remain outside design/mechanics layers.
"""

from collections.abc import Iterable
from pathlib import Path

import pytest
from openpyxl import load_workbook


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
MATERIALS_PATH = REPOSITORY_ROOT / "data" / "catalogs" / "materials_catalog.xlsx"
SECTIONS_PATH = REPOSITORY_ROOT / "data" / "catalogs" / "sections_catalog.xlsx"
PROJECT_PATH = REPOSITORY_ROOT / "projects" / "PRJ_001"
MEMBERS_PATH = PROJECT_PATH / "members.xlsx"
ETABS_PATH = PROJECT_PATH / "ETABS_results.xlsx"
PROJECT_YAML_PATH = PROJECT_PATH / "project.yaml"

APPROVED_CONTRACT_PATHS = (
    MATERIALS_PATH,
    SECTIONS_PATH,
    MEMBERS_PATH,
    ETABS_PATH,
    PROJECT_YAML_PATH,
)


def _headers(path: Path, sheet: str, *, row_number: int = 1) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        values = next(
            workbook[sheet].iter_rows(
                min_row=row_number,
                max_row=row_number,
                values_only=True,
            )
        )
        return {str(value) for value in values if value is not None}
    finally:
        workbook.close()


def _assert_required_sheets(path: Path, expected: Iterable[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = set(expected) - set(workbook.sheetnames)
        assert not missing, f"{path.name} is missing sheets: {sorted(missing)}"
    finally:
        workbook.close()


def _assert_columns(
    path: Path,
    sheet: str,
    expected: Iterable[str],
    *,
    row_number: int = 1,
) -> None:
    missing = set(expected) - _headers(path, sheet, row_number=row_number)
    assert not missing, f"{path.name}:{sheet} is missing columns: {sorted(missing)}"


def _assert_self_declared_required_columns(path: Path) -> None:
    """Check every field marked required by the workbook's Schema sheet."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        schema_rows = workbook["Schema"].iter_rows(values_only=True)
        schema_headers = next(schema_rows)
        positions = {name: index for index, name in enumerate(schema_headers)}
        assert {"Sheet", "Field", "Required"} <= positions.keys()

        headers_by_sheet: dict[str, set[str]] = {}
        for row in schema_rows:
            if row[positions["Required"]] != "YES":
                continue
            sheet = str(row[positions["Sheet"]])
            field = str(row[positions["Field"]])
            assert sheet in workbook.sheetnames, f"Schema references missing sheet {sheet}"
            if sheet not in headers_by_sheet:
                values = next(workbook[sheet].iter_rows(max_row=1, values_only=True))
                headers_by_sheet[sheet] = {
                    str(value) for value in values if value is not None
                }
            assert field in headers_by_sheet[sheet], (
                f"{path.name}:{sheet} is missing required schema field {field}"
            )
    finally:
        workbook.close()


def test_all_five_approved_input_contracts_exist() -> None:
    missing = [
        str(path.relative_to(REPOSITORY_ROOT))
        for path in APPROVED_CONTRACT_PATHS
        if not path.is_file()
    ]
    assert not missing, f"Missing approved input contracts: {missing}"


@pytest.mark.parametrize(
    ("path", "sheets"),
    (
        (MATERIALS_PATH, {"Materials", "Sources", "Metadata", "Schema"}),
        (
            SECTIONS_PATH,
            {
                "Sections",
                "Geometry",
                "Properties",
                "AISI_Dimensions",
                "Sources",
                "Metadata",
                "Schema",
            },
        ),
        (MEMBERS_PATH, {"Members", "ETABS_Mapping", "Metadata", "Schema"}),
    ),
)
def test_versioned_workbooks_contain_required_sheets(
    path: Path, sheets: set[str]
) -> None:
    _assert_required_sheets(path, sheets)


@pytest.mark.parametrize("path", (MATERIALS_PATH, SECTIONS_PATH, MEMBERS_PATH))
def test_versioned_workbooks_satisfy_their_required_schema_fields(path: Path) -> None:
    _assert_self_declared_required_columns(path)


def test_materials_catalog_core_columns() -> None:
    _assert_columns(
        MATERIALS_PATH,
        "Materials",
        {
            "material_id",
            "designation",
            "specification",
            "grade",
            "Fy_MPa",
            "Fu_MPa",
            "E_MPa",
            "nu",
            "source_id",
            "active",
        },
    )


def test_sections_catalog_core_columns() -> None:
    _assert_columns(
        SECTIONS_PATH,
        "Sections",
        {"section_id", "designation", "family", "geometry_id", "source_id", "active"},
    )
    _assert_columns(
        SECTIONS_PATH,
        "Geometry",
        {"geometry_id", "section_type", "H_mm", "B1_mm", "t_mm", "Ri_mm"},
    )
    _assert_columns(
        SECTIONS_PATH,
        "Properties",
        {"section_id", "A_mm2", "Ix_mm4", "Iy_mm4", "J_mm4", "Cw_mm6"},
    )
    _assert_columns(
        SECTIONS_PATH,
        "AISI_Dimensions",
        {
            "geometry_id",
            "standard_id",
            "standard_edition",
            "web_flat_width_mm",
            "flange_1_flat_width_mm",
            "flange_2_flat_width_mm",
            "web_out_to_out_depth_mm",
            "flange_1_out_to_out_width_mm",
            "flange_2_out_to_out_width_mm",
            "lip_1_flat_width_mm",
            "lip_2_flat_width_mm",
            "lip_1_out_to_out_width_mm",
            "lip_2_out_to_out_width_mm",
            "lip_1_overall_depth_mm",
            "lip_2_overall_depth_mm",
            "source_id",
        },
    )


def test_members_workbook_core_columns() -> None:
    _assert_columns(
        MEMBERS_PATH,
        "Members",
        {
            "case_id",
            "label",
            "member_type",
            "section_id",
            "material_id",
            "L_mm",
            "length_definition",
            "active",
        },
    )
    _assert_columns(
        MEMBERS_PATH,
        "ETABS_Mapping",
        {
            "case_id",
            "etabs_unique_name",
            "etabs_story",
            "etabs_beam",
            "mapping_enabled",
        },
    )


def test_etabs_native_sheets_and_force_columns_are_preserved() -> None:
    workbook = load_workbook(ETABS_PATH, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Program Control", "Element Forces - Beams"]
    finally:
        workbook.close()

    _assert_columns(
        ETABS_PATH,
        "Program Control",
        {
            "ProgramName",
            "Version",
            "CurrUnits",
        },
        row_number=2,
    )
    # Native ETABS tables place the field names on row 2, after a table title.
    force_headers = _headers(ETABS_PATH, "Element Forces - Beams", row_number=2)
    expected = {
        "Story",
        "Beam",
        "Unique Name",
        "Output Case",
        "Case Type",
        "Step Type",
        "Station",
        "P",
        "V2",
        "V3",
        "T",
        "M2",
        "M3",
        "Element",
        "Elem Station",
        "Location",
    }
    assert expected <= force_headers
