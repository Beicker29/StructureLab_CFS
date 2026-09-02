"""M2 catalog loader tests using approved files and temporary invalid copies."""

from collections.abc import Callable
from hashlib import sha256
from pathlib import Path
from shutil import copy2

import pytest
from openpyxl import load_workbook

from cfs_design.catalogs import load_material_catalog, load_section_catalog
from cfs_design.core.exceptions import CatalogError, SchemaError
from cfs_design.domain import (
    GeometryConvention,
    S100_24_STANDARD_EDITION,
    S100_24_STANDARD_ID,
    SectionFamily,
    StandardSectionDimensions,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
MATERIALS_SOURCE = REPOSITORY_ROOT / "data" / "catalogs" / "materials_catalog.xlsx"
SECTIONS_SOURCE = REPOSITORY_ROOT / "data" / "catalogs" / "sections_catalog.xlsx"


@pytest.fixture
def materials_copy(tmp_path: Path) -> Path:
    target = tmp_path / "materials_catalog.xlsx"
    copy2(MATERIALS_SOURCE, target)
    return target


@pytest.fixture
def sections_copy(tmp_path: Path) -> Path:
    target = tmp_path / "sections_catalog.xlsx"
    copy2(SECTIONS_SOURCE, target)
    return target


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _column(worksheet: object, header: str) -> int:
    for cell in worksheet[1]:  # type: ignore[index]
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Test fixture has no column {header}")


def _set_cell(
    path: Path,
    worksheet_name: str,
    row_number: int,
    header: str,
    value: object,
) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook[worksheet_name]  # type: ignore[index]
        worksheet.cell(row_number, _column(worksheet, header), value)

    _modify(path, operation)


def _append_copy(
    path: Path,
    worksheet_name: str,
    source_row: int = 2,
    replacements: dict[str, object] | None = None,
) -> None:
    replacements = replacements or {}

    def operation(workbook: object) -> None:
        worksheet = workbook[worksheet_name]  # type: ignore[index]
        values = [cell.value for cell in worksheet[source_row]]
        for header, value in replacements.items():
            values[_column(worksheet, header) - 1] = value
        worksheet.append(values)

    _modify(path, operation)


def _append_synthetic_aisi_dimensions(
    path: Path,
    replacements: dict[str, object] | None = None,
) -> None:
    values: dict[str, object] = {
        "geometry_id": "EX_GEO_C200",
        "standard_id": S100_24_STANDARD_ID,
        "standard_edition": S100_24_STANDARD_EDITION,
        "web_flat_width_mm": 196.0,
        "flange_1_flat_width_mm": 66.0,
        "flange_2_flat_width_mm": 66.0,
        "web_out_to_out_depth_mm": 204.0,
        "flange_1_out_to_out_width_mm": 72.0,
        "flange_2_out_to_out_width_mm": 72.0,
        "lip_1_flat_width_mm": 16.0,
        "lip_2_flat_width_mm": 16.0,
        "lip_1_out_to_out_width_mm": 20.0,
        "lip_2_out_to_out_width_mm": 20.0,
        "lip_1_overall_depth_mm": 20.0,
        "lip_2_overall_depth_mm": 20.0,
        "source_id": "EX_SRC_SEC",
        "notes": "SYNTHETIC_TEST_DATA",
    }
    values.update(replacements or {})

    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Dimensions"]  # type: ignore[index]
        worksheet.append([values.get(cell.value) for cell in worksheet[1]])

    _modify(path, operation)


def _make_legacy_section_catalog(path: Path) -> None:
    def operation(workbook: object) -> None:
        workbook.remove(workbook["AISI_Dimensions"])  # type: ignore[attr-defined,index]
        metadata = workbook["Metadata"]  # type: ignore[index]
        headers = {_cell.value: _cell.column for _cell in metadata[1]}
        for row_number in range(2, metadata.max_row + 1):
            if metadata.cell(row_number, headers["Field"]).value == "schema_version":
                metadata.cell(row_number, headers["Value"]).value = "0.1.0"
                return
        raise AssertionError("Metadata has no schema_version row")

    _modify(path, operation)


def test_approved_material_catalog_loads_two_inactive_materials() -> None:
    catalog = load_material_catalog(MATERIALS_SOURCE)

    assert catalog.metadata.schema_version == "0.2.0"
    assert catalog.metadata.canonical_units == "SI"
    assert dict(catalog.metadata.additional_fields)["created"] == "2026-08-31"
    assert len(catalog.sources) == 1
    assert catalog.sources[0].source_id == "EX_SRC_MAT"
    assert tuple(material.material_id for material in catalog.materials) == (
        "EX_MAT_G33",
        "EX_MAT_G50",
    )
    assert all(not material.active for material in catalog.materials)
    assert catalog.active_materials == ()
    assert catalog.material_qualifications == ()


def test_material_catalog_records_absolute_path_and_sha256() -> None:
    catalog = load_material_catalog(MATERIALS_SOURCE)

    assert catalog.metadata.source_path == MATERIALS_SOURCE.resolve()
    assert catalog.metadata.file_sha256 == sha256(MATERIALS_SOURCE.read_bytes()).hexdigest()


def test_missing_catalog_file_is_rejected(tmp_path: Path) -> None:
    with pytest.raises(CatalogError, match="does not exist"):
        load_material_catalog(tmp_path / "missing.xlsx")


@pytest.mark.parametrize(
    "sheet_name",
    (
        "Materials",
        "AISI_Material_Qualification",
        "Sources",
        "Metadata",
        "Schema",
    ),
)
def test_missing_material_workbook_sheet_is_rejected(
    materials_copy: Path,
    sheet_name: str,
) -> None:
    _modify(materials_copy, lambda workbook: workbook.remove(workbook[sheet_name]))  # type: ignore[attr-defined,index]

    with pytest.raises(SchemaError, match=sheet_name):
        load_material_catalog(materials_copy)


def test_missing_material_required_column_is_rejected(materials_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Materials"]  # type: ignore[index]
        worksheet.delete_cols(_column(worksheet, "Fy_MPa"))

    _modify(materials_copy, operation)
    with pytest.raises(SchemaError, match="Fy_MPa"):
        load_material_catalog(materials_copy)


def test_unsupported_material_schema_version_is_rejected(materials_copy: Path) -> None:
    _set_cell(materials_copy, "Metadata", 3, "Value", "9.9.9")

    with pytest.raises(SchemaError, match="unsupported schema_version"):
        load_material_catalog(materials_copy)


def test_duplicate_material_id_is_rejected(materials_copy: Path) -> None:
    _append_copy(materials_copy, "Materials")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_material_catalog(materials_copy)


def test_duplicate_material_source_id_is_rejected(materials_copy: Path) -> None:
    _append_copy(materials_copy, "Sources")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_material_catalog(materials_copy)


def test_missing_material_source_reference_is_rejected(materials_copy: Path) -> None:
    _set_cell(materials_copy, "Materials", 2, "source_id", "UNKNOWN")

    with pytest.raises(CatalogError, match="Unknown reference"):
        load_material_catalog(materials_copy)


def test_invalid_material_numeric_cell_has_context(materials_copy: Path) -> None:
    _set_cell(materials_copy, "Materials", 2, "Fy_MPa", "abc")

    with pytest.raises(CatalogError) as captured:
        load_material_catalog(materials_copy)
    message = str(captured.value)
    assert "materials_catalog.xlsx" in message
    assert "Worksheet: Materials" in message
    assert "Row: 2" in message
    assert "Field: Fy_MPa" in message
    assert "Invalid numeric value" in message


@pytest.mark.parametrize("value", ("yes", 1, 0, ""))
def test_invalid_material_boolean_is_rejected(
    materials_copy: Path,
    value: object,
) -> None:
    _set_cell(materials_copy, "Materials", 2, "active", value)

    with pytest.raises(CatalogError, match="Invalid boolean value"):
        load_material_catalog(materials_copy)


def test_controlled_text_boolean_is_supported(materials_copy: Path) -> None:
    _set_cell(materials_copy, "Materials", 2, "active", "TRUE")

    catalog = load_material_catalog(materials_copy)
    assert catalog.materials[0].active is True
    assert tuple(item.material_id for item in catalog.active_materials) == (
        "EX_MAT_G33",
    )


def test_blank_material_rows_are_ignored(materials_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook["Materials"].insert_rows(3)  # type: ignore[index]

    _modify(materials_copy, operation)
    catalog = load_material_catalog(materials_copy)
    assert len(catalog.materials) == 2


def test_partially_populated_material_row_is_rejected(materials_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Materials"]  # type: ignore[index]
        values: list[object | None] = [None] * worksheet.max_column
        values[_column(worksheet, "material_id") - 1] = "PARTIAL"
        worksheet.append(values)

    _modify(materials_copy, operation)
    with pytest.raises(CatalogError, match="Invalid required text value"):
        load_material_catalog(materials_copy)


def test_unresolved_required_formula_is_rejected(materials_copy: Path) -> None:
    _set_cell(materials_copy, "Materials", 2, "Fy_MPa", "=200+30")

    with pytest.raises(CatalogError, match="Formula cell has no usable cached value"):
        load_material_catalog(materials_copy)


def test_approved_section_catalog_loads_two_resolved_inactive_sections() -> None:
    catalog = load_section_catalog(SECTIONS_SOURCE)

    assert catalog.metadata.schema_version == "0.2.0"
    assert len(catalog.sections) == 2
    assert len(catalog.geometries) == 2
    assert len(catalog.properties) == 2
    assert len(catalog.resolved_sections) == 2
    assert catalog.standard_dimensions == ()
    assert all(not section.catalog_section.active for section in catalog.resolved_sections)
    assert catalog.active_sections == ()


def test_legacy_section_catalog_0_1_0_still_loads_without_dimensions(
    sections_copy: Path,
) -> None:
    _make_legacy_section_catalog(sections_copy)

    catalog = load_section_catalog(sections_copy)

    assert catalog.metadata.schema_version == "0.1.0"
    assert catalog.standard_dimensions == ()
    assert all(not item.standard_dimensions for item in catalog.resolved_sections)


def test_valid_synthetic_aisi_dimension_record_is_typed_and_resolved(
    sections_copy: Path,
) -> None:
    _append_synthetic_aisi_dimensions(sections_copy)

    catalog = load_section_catalog(sections_copy)
    dimensions = catalog.get_standard_dimensions(
        "EX_GEO_C200",
        S100_24_STANDARD_ID,
        S100_24_STANDARD_EDITION,
    )

    assert isinstance(dimensions, StandardSectionDimensions)
    assert dimensions.source_id == "EX_SRC_SEC"
    assert dimensions.notes == "SYNTHETIC_TEST_DATA"
    assert dimensions.flange_1_flat_width_mm == 66.0
    assert catalog.get_section(
        "EX_SEC_C200_70_20_2"
    ).standard_dimensions == (dimensions,)


def test_midline_geometry_does_not_create_aisi_dimensions() -> None:
    catalog = load_section_catalog(SECTIONS_SOURCE)

    assert catalog.standard_dimensions == ()
    assert all(item.standard_dimensions == () for item in catalog.resolved_sections)
    with pytest.raises(CatalogError, match="Unknown standard dimensions"):
        catalog.get_standard_dimensions(
            "EX_GEO_C200",
            S100_24_STANDARD_ID,
            S100_24_STANDARD_EDITION,
        )


def test_section_catalog_records_absolute_path_and_sha256() -> None:
    catalog = load_section_catalog(SECTIONS_SOURCE)

    assert catalog.metadata.source_path == SECTIONS_SOURCE.resolve()
    assert catalog.metadata.file_sha256 == sha256(SECTIONS_SOURCE.read_bytes()).hexdigest()


def test_active_section_view_filters_without_discarding_inactive_rows(
    sections_copy: Path,
) -> None:
    _set_cell(sections_copy, "Sections", 2, "active", True)

    catalog = load_section_catalog(sections_copy)
    assert len(catalog.sections) == 2
    assert tuple(
        section.catalog_section.section_id for section in catalog.active_sections
    ) == ("EX_SEC_C200_70_20_2",)


def test_approved_resolved_section_preserves_enums_and_optional_values() -> None:
    catalog = load_section_catalog(SECTIONS_SOURCE)
    section = catalog.get_section("EX_SEC_C200_70_20_2")

    assert section.catalog_section.family is SectionFamily.C_LIPPED
    assert section.geometry.geometry_convention is GeometryConvention.MIDLINE
    assert section.properties.cw_mm6 is None
    assert section.properties.x0_mm is None
    assert section.properties.y0_mm is None


@pytest.mark.parametrize(
    "sheet_name",
    (
        "Sections",
        "Geometry",
        "Properties",
        "AISI_Dimensions",
        "Sources",
        "Metadata",
        "Schema",
    ),
)
def test_missing_section_workbook_sheet_is_rejected(
    sections_copy: Path,
    sheet_name: str,
) -> None:
    _modify(sections_copy, lambda workbook: workbook.remove(workbook[sheet_name]))  # type: ignore[attr-defined,index]

    with pytest.raises(SchemaError, match=sheet_name):
        load_section_catalog(sections_copy)


def test_missing_section_required_column_is_rejected(sections_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["Properties"]  # type: ignore[index]
        worksheet.delete_cols(_column(worksheet, "Ix_mm4"))

    _modify(sections_copy, operation)
    with pytest.raises(SchemaError, match="Ix_mm4"):
        load_section_catalog(sections_copy)


def test_missing_aisi_dimension_required_column_is_rejected(
    sections_copy: Path,
) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Dimensions"]  # type: ignore[index]
        worksheet.delete_cols(_column(worksheet, "standard_edition"))

    _modify(sections_copy, operation)
    with pytest.raises(SchemaError, match="standard_edition"):
        load_section_catalog(sections_copy)


def test_duplicate_aisi_dimension_composite_key_is_rejected(
    sections_copy: Path,
) -> None:
    _append_synthetic_aisi_dimensions(sections_copy)
    _append_synthetic_aisi_dimensions(sections_copy)

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_section_catalog(sections_copy)


@pytest.mark.parametrize(
    ("field", "value", "target"),
    (
        ("geometry_id", "UNKNOWN", "Geometry.geometry_id"),
        ("source_id", "UNKNOWN", "Sources.source_id"),
    ),
)
def test_aisi_dimension_unknown_reference_is_rejected(
    sections_copy: Path,
    field: str,
    value: object,
    target: str,
) -> None:
    _append_synthetic_aisi_dimensions(sections_copy, {field: value})

    with pytest.raises(CatalogError, match=target):
        load_section_catalog(sections_copy)


@pytest.mark.parametrize("value", ("abc", -1.0, float("inf")))
def test_invalid_aisi_dimension_numeric_value_is_rejected(
    sections_copy: Path,
    value: object,
) -> None:
    _append_synthetic_aisi_dimensions(
        sections_copy,
        {"web_flat_width_mm": value},
    )

    with pytest.raises(CatalogError, match="Invalid"):
        load_section_catalog(sections_copy)


def test_partial_aisi_dimension_row_is_rejected(sections_copy: Path) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Dimensions"]  # type: ignore[index]
        values: list[object | None] = [None] * worksheet.max_column
        values[_column(worksheet, "geometry_id") - 1] = "EX_GEO_C200"
        worksheet.append(values)

    _modify(sections_copy, operation)
    with pytest.raises(CatalogError, match="Invalid required text value"):
        load_section_catalog(sections_copy)


def test_blank_aisi_dimension_row_is_ignored(sections_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook["AISI_Dimensions"].append(  # type: ignore[index]
            [None] * workbook["AISI_Dimensions"].max_column  # type: ignore[index]
        )

    _modify(sections_copy, operation)
    assert load_section_catalog(sections_copy).standard_dimensions == ()


def test_unsupported_aisi_dimension_edition_is_rejected(
    sections_copy: Path,
) -> None:
    _append_synthetic_aisi_dimensions(
        sections_copy,
        {"standard_edition": 2016},
    )

    with pytest.raises(CatalogError, match="Unsupported standard-specific"):
        load_section_catalog(sections_copy)


def test_incomplete_lipped_dimension_group_is_rejected(
    sections_copy: Path,
) -> None:
    _append_synthetic_aisi_dimensions(
        sections_copy,
        {"lip_2_overall_depth_mm": None},
    )

    with pytest.raises(CatalogError, match="complete set"):
        load_section_catalog(sections_copy)


def test_duplicate_section_id_is_rejected(sections_copy: Path) -> None:
    _append_copy(sections_copy, "Sections")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_section_catalog(sections_copy)


def test_duplicate_geometry_id_is_rejected(sections_copy: Path) -> None:
    _append_copy(sections_copy, "Geometry")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_section_catalog(sections_copy)


def test_duplicate_property_section_id_is_rejected(sections_copy: Path) -> None:
    _append_copy(sections_copy, "Properties")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_section_catalog(sections_copy)


def test_duplicate_section_source_id_is_rejected(sections_copy: Path) -> None:
    _append_copy(sections_copy, "Sources")

    with pytest.raises(CatalogError, match="Duplicate ID"):
        load_section_catalog(sections_copy)


def test_missing_geometry_reference_is_rejected(sections_copy: Path) -> None:
    _set_cell(sections_copy, "Sections", 2, "geometry_id", "UNKNOWN")

    with pytest.raises(CatalogError, match="Unknown reference"):
        load_section_catalog(sections_copy)


def test_missing_property_row_is_rejected(sections_copy: Path) -> None:
    def operation(workbook: object) -> None:
        workbook["Properties"].delete_rows(2)  # type: ignore[index]

    _modify(sections_copy, operation)
    with pytest.raises(CatalogError, match="Properties.section_id"):
        load_section_catalog(sections_copy)


@pytest.mark.parametrize(
    ("worksheet_name", "field"),
    (("Sections", "source_id"), ("Properties", "source_id")),
)
def test_missing_section_source_reference_is_rejected(
    sections_copy: Path,
    worksheet_name: str,
    field: str,
) -> None:
    _set_cell(sections_copy, worksheet_name, 2, field, "UNKNOWN")

    with pytest.raises(CatalogError, match="Sources.source_id"):
        load_section_catalog(sections_copy)


def test_family_and_section_type_mismatch_is_rejected(sections_copy: Path) -> None:
    _set_cell(sections_copy, "Geometry", 2, "section_type", "C_UNLIPPED")

    with pytest.raises(CatalogError, match="family"):
        load_section_catalog(sections_copy)


@pytest.mark.parametrize(
    ("worksheet_name", "field"),
    (("Sections", "family"), ("Geometry", "geometry_convention")),
)
def test_unknown_section_enum_value_is_rejected(
    sections_copy: Path,
    worksheet_name: str,
    field: str,
) -> None:
    _set_cell(sections_copy, worksheet_name, 2, field, "UNKNOWN_ENUM")

    with pytest.raises(CatalogError, match="Unknown"):
        load_section_catalog(sections_copy)


def test_orphan_geometry_is_rejected(sections_copy: Path) -> None:
    _append_copy(
        sections_copy,
        "Geometry",
        replacements={"geometry_id": "ORPHAN_GEO"},
    )

    with pytest.raises(CatalogError, match="Orphan geometry"):
        load_section_catalog(sections_copy)


def test_orphan_property_is_rejected(sections_copy: Path) -> None:
    _append_copy(
        sections_copy,
        "Properties",
        replacements={"section_id": "ORPHAN_SECTION"},
    )

    with pytest.raises(CatalogError, match="Orphan property"):
        load_section_catalog(sections_copy)
