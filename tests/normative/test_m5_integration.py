"""M5-shaped temporary project exercised through the M7 eligibility gate."""

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

from cfs_design.domain import DesignMethod
from cfs_design.normative import (
    DesignAction,
    SoftwareSupportStatus,
    evaluate_design_eligibility,
)
from cfs_design.results import ApplicabilityStatus
from cfs_design.workflows.project import resolve_project


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]


def _copy_approved_project(tmp_path: Path) -> Path:
    root = tmp_path / "repository"
    catalogs = root / "data" / "catalogs"
    project = root / "projects" / "PRJ_001"
    catalogs.mkdir(parents=True)
    project.mkdir(parents=True)
    for name in ("materials_catalog.xlsx", "sections_catalog.xlsx"):
        copy2(REPOSITORY_ROOT / "data" / "catalogs" / name, catalogs / name)
    for name in ("members.xlsx", "ETABS_results.xlsx", "project.yaml"):
        copy2(REPOSITORY_ROOT / "projects" / "PRJ_001" / name, project / name)
    return root


def _activate_row(
    path: Path,
    sheet_name: str,
    id_header: str,
    identifier: str,
    field: str,
) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, headers[id_header]).value == identifier:
            worksheet.cell(row, headers[field]).value = True
            workbook.save(path)
            workbook.close()
            return
    workbook.close()
    raise AssertionError(f"No {identifier} in {sheet_name}")


def _activate_example(root: Path) -> None:
    _activate_row(
        root / "data" / "catalogs" / "materials_catalog.xlsx",
        "Materials",
        "material_id",
        "EX_MAT_G50",
        "active",
    )
    _activate_row(
        root / "data" / "catalogs" / "sections_catalog.xlsx",
        "Sections",
        "section_id",
        "EX_SEC_C200_70_20_2",
        "active",
    )
    members = root / "projects" / "PRJ_001" / "members.xlsx"
    _activate_row(
        members,
        "Members",
        "case_id",
        "EX_BEAM_001",
        "active",
    )
    _activate_row(
        members,
        "ETABS_Mapping",
        "case_id",
        "EX_BEAM_001",
        "mapping_enabled",
    )


def test_real_resolved_member_stays_indeterminate_when_facts_are_absent(
    tmp_path: Path,
) -> None:
    root = _copy_approved_project(tmp_path)
    _activate_example(root)
    resolved = resolve_project(
        root / "projects" / "PRJ_001" / "project.yaml",
        repository_root=root,
    )
    member = resolved.active_resolved_members[0]

    eligibility = evaluate_design_eligibility(
        member,
        resolved.design_context,
        DesignMethod.EWM,
        DesignAction.STRONG_AXIS_FLEXURE,
    )

    assert eligibility.normative.status is ApplicabilityStatus.INDETERMINATE
    assert eligibility.software.status is SoftwareSupportStatus.SUPPORTED
    assert eligibility.executable is False
    assert any(
        item.code == "AISI_B4_DIMENSION_BASIS_UNAVAILABLE"
        for item in eligibility.normative.diagnostics
    )
