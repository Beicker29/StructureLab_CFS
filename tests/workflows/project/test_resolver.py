"""M5 project resolver integration and QA policy tests."""

from collections.abc import Callable
from dataclasses import replace
from hashlib import sha256
from importlib import import_module
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from cfs_design.core.exceptions import ConfigurationError, ValidationError
from cfs_design.domain import SectionDemandSet
from cfs_design.mechanics.sections import (
    AdvancedSectionProperties,
    ComputedSectionProperties,
    VerificationStatus,
)
from cfs_design.workflows.project import DiagnosticSeverity, resolve_project


REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
APPROVED_PROJECT = REPOSITORY_ROOT / "projects" / "PRJ_001" / "project.yaml"


def _project_yaml(root: Path) -> Path:
    return root / "projects" / "PRJ_001" / "project.yaml"


def _workbook(root: Path, name: str) -> Path:
    if name in {"materials_catalog.xlsx", "sections_catalog.xlsx"}:
        return root / "data" / "catalogs" / name
    return root / "projects" / "PRJ_001" / name


def _modify_workbook(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _headers(worksheet: object, row: int = 1) -> dict[object, int]:
    return {cell.value: cell.column for cell in worksheet[row]}  # type: ignore[index]


def _set_by_id(
    path: Path,
    sheet_name: str,
    id_header: str,
    identifier: str,
    **updates: object,
) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook[sheet_name]  # type: ignore[index]
        headers = _headers(worksheet)
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, headers[id_header]).value == identifier:
                for header, value in updates.items():
                    worksheet.cell(row, headers[header]).value = value
                return
        raise AssertionError(f"No {identifier} in {sheet_name}")

    _modify_workbook(path, operation)


def _edit_yaml(root: Path, operation: Callable[[dict[str, object]], None]) -> None:
    path = _project_yaml(root)
    document = yaml.safe_load(path.read_text(encoding="utf-8"))
    operation(document)
    path.write_text(yaml.safe_dump(document, sort_keys=False), encoding="utf-8")


def _activate_example(root: Path) -> None:
    _set_by_id(
        _workbook(root, "materials_catalog.xlsx"),
        "Materials",
        "material_id",
        "EX_MAT_G50",
        active=True,
    )
    _set_by_id(
        _workbook(root, "sections_catalog.xlsx"),
        "Sections",
        "section_id",
        "EX_SEC_C200_70_20_2",
        active=True,
    )
    _set_by_id(
        _workbook(root, "members.xlsx"),
        "Members",
        "case_id",
        "EX_BEAM_001",
        active=True,
    )
    _set_by_id(
        _workbook(root, "members.xlsx"),
        "ETABS_Mapping",
        "case_id",
        "EX_BEAM_001",
        mapping_enabled=True,
    )


def _upgrade_members_to_schema_02(root: Path) -> None:
    def operation(workbook: object) -> None:
        metadata = workbook["Metadata"]  # type: ignore[index]
        for row in range(2, metadata.max_row + 1):
            if metadata.cell(row, 1).value == "schema_version":
                metadata.cell(row, 2).value = "0.2.0"
                break
        members = workbook["Members"]  # type: ignore[index]
        headers = _headers(members)
        if "distortional_unbraced_length_mm" not in headers:
            members.cell(1, members.max_column + 1).value = (
                "distortional_unbraced_length_mm"
            )
        headers = _headers(members)
        if "distortional_restraint_source" not in headers:
            members.cell(1, members.max_column + 1).value = (
                "distortional_restraint_source"
            )

    _modify_workbook(_workbook(root, "members.xlsx"), operation)


def test_approved_inactive_project_loads_without_forcing_execution() -> None:
    resolved = resolve_project(APPROVED_PROJECT, repository_root=REPOSITORY_ROOT)

    assert len(resolved.all_member_cases) == 2
    assert len(resolved.inactive_member_cases) == 2
    assert resolved.active_resolved_members == ()
    assert resolved.section_verification_results == ()
    assert len(resolved.unmapped_etabs_rows) == 24
    assert resolved.etabs_import.metadata.program_version == "18.1.1"
    assert any(item.code == "ETABS_IMPORT_WARNING" for item in resolved.diagnostics)


def test_active_example_resolves_catalogs_demands_verification_and_provenance(
    project_root: Path,
) -> None:
    _activate_example(project_root)
    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )

    assert len(resolved.active_resolved_members) == 1
    member = resolved.active_resolved_members[0]
    assert member.member.case_id == "EX_BEAM_001"
    assert member.section.catalog_section.section_id == "EX_SEC_C200_70_20_2"
    assert member.material.material_id == "EX_MAT_G50"
    assert isinstance(member.demands, SectionDemandSet)
    assert member.source_demands is not None
    assert tuple(item.combination_id for item in member.demands.combinations) == (
        "DERX",
        "DERY",
    )
    assert tuple(len(item.points) for item in member.demands.combinations) == (12, 12)
    source_first = member.source_demands.combinations[0].points[0]
    section_first = member.demands.combinations[0].points[0]
    assert section_first.mx_nmm == source_first.m2_nmm
    assert section_first.my_nmm == source_first.m3_nmm
    assert section_first.source_point_id == source_first.point_id
    assert len(resolved.section_verification_results) == 1
    assert resolved.section_verification_results[0].overall_status is VerificationStatus.PASS
    mechanics = resolved.get_section_mechanics("EX_SEC_C200_70_20_2")
    assert resolved.require_design_mechanics("EX_SEC_C200_70_20_2") is mechanics
    assert isinstance(mechanics.gross, ComputedSectionProperties)
    assert isinstance(mechanics.advanced, AdvancedSectionProperties)
    assert mechanics.verification is resolved.section_verification_results[0]
    assert mechanics.design_use_permitted
    assert mechanics.gross.section_id == member.section.catalog_section.section_id
    assert mechanics.advanced.section_id == member.section.catalog_section.section_id
    assert mechanics.gross is not member.section.properties
    assert {
        item.code for item in resolved.diagnostics
    } >= {"CATALOG_EXTENDED_NOT_CHECKED"}
    assert resolved.unmapped_etabs_rows == ()
    provenance = resolved.provenance
    assert provenance.project_yaml_sha256 == sha256(_project_yaml(project_root).read_bytes()).hexdigest()
    assert provenance.members_sha256 == sha256(_workbook(project_root, "members.xlsx").read_bytes()).hexdigest()
    assert provenance.materials_catalog_sha256 == sha256(_workbook(project_root, "materials_catalog.xlsx").read_bytes()).hexdigest()
    assert provenance.sections_catalog_sha256 == sha256(_workbook(project_root, "sections_catalog.xlsx").read_bytes()).hexdigest()
    assert provenance.etabs_sha256 == sha256(_workbook(project_root, "ETABS_results.xlsx").read_bytes()).hexdigest()
    assert provenance.etabs_program_version == "18.1.1"
    assert not resolved.project_config.outputs.resolved_root.exists()


def test_explicit_lm_survives_members_loader_and_m5_resolution(
    project_root: Path,
) -> None:
    _upgrade_members_to_schema_02(project_root)
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "members.xlsx"),
        "Members",
        "case_id",
        "EX_BEAM_001",
        distortional_unbraced_length_mm=1800.0,
        distortional_restraint_source="Structural restraint schedule R-01.",
    )

    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )
    restraints = resolved.active_resolved_members[0].member.restraints

    assert restraints.distortional_unbraced_length_mm == 1800.0
    assert restraints.distortional_restraint_source == (
        "Structural restraint schedule R-01."
    )


def test_all_stations_and_before_after_survive_resolution(project_root: Path) -> None:
    _activate_example(project_root)
    member = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    ).active_resolved_members[0]

    assert member.source_demands is not None
    source_count = sum(len(item.points) for item in member.source_demands.combinations)
    section_count = sum(len(item.points) for item in member.demands.combinations)
    assert source_count == section_count == 24
    for combination in member.demands.combinations:
        discontinuity = [
            point for point in combination.points if point.station_mm == 2750.0
        ]
        assert len(discontinuity) == 2
        assert {point.location for point in discontinuity} == {"Before", "After"}


@pytest.mark.parametrize(
    ("field", "value", "match"),
    (
        ("section_id", "UNKNOWN_SECTION", "unknown section_id"),
        ("material_id", "UNKNOWN_MATERIAL", "unknown material_id"),
    ),
)
def test_unknown_active_catalog_reference_is_fatal(
    project_root: Path,
    field: str,
    value: str,
    match: str,
) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "members.xlsx"),
        "Members",
        "case_id",
        "EX_BEAM_001",
        **{field: value},
    )
    with pytest.raises(ConfigurationError, match=match):
        resolve_project(_project_yaml(project_root), repository_root=project_root)


@pytest.mark.parametrize("catalog", ("section", "material"))
def test_active_member_cannot_use_inactive_catalog_record(
    project_root: Path,
    catalog: str,
) -> None:
    _activate_example(project_root)
    if catalog == "section":
        _set_by_id(
            _workbook(project_root, "sections_catalog.xlsx"),
            "Sections",
            "section_id",
            "EX_SEC_C200_70_20_2",
            active=False,
        )
    else:
        _set_by_id(
            _workbook(project_root, "materials_catalog.xlsx"),
            "Materials",
            "material_id",
            "EX_MAT_G50",
            active=False,
        )
    with pytest.raises(ConfigurationError, match=f"inactive {catalog}"):
        resolve_project(_project_yaml(project_root), repository_root=project_root)


def test_active_member_requires_enabled_mapping(project_root: Path) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "members.xlsx"),
        "ETABS_Mapping",
        "case_id",
        "EX_BEAM_001",
        mapping_enabled=False,
    )
    with pytest.raises(ConfigurationError, match="no mapped usable ETABS DemandSet"):
        resolve_project(_project_yaml(project_root), repository_root=project_root)


def test_active_enabled_mapping_with_no_rows_is_rejected(project_root: Path) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "members.xlsx"),
        "ETABS_Mapping",
        "case_id",
        "EX_BEAM_001",
        etabs_unique_name="UNKNOWN",
        etabs_story="UNKNOWN",
        etabs_beam="UNKNOWN",
    )
    with pytest.raises(ConfigurationError, match="no mapped usable ETABS DemandSet"):
        resolve_project(_project_yaml(project_root), repository_root=project_root)


def test_extra_unmapped_etabs_rows_are_diagnostic_not_fatal(
    project_root: Path,
) -> None:
    _activate_example(project_root)

    def alter_one_row(workbook: object) -> None:
        worksheet = workbook["Element Forces - Beams"]  # type: ignore[index]
        headers = _headers(worksheet, row=2)
        worksheet.cell(4, headers["Unique Name"]).value = "OUTSIDE"
        worksheet.cell(4, headers["Story"]).value = "OUTSIDE"
        worksheet.cell(4, headers["Beam"]).value = "OUTSIDE"

    _modify_workbook(_workbook(project_root, "ETABS_results.xlsx"), alter_one_row)
    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )
    assert len(resolved.active_resolved_members) == 1
    assert len(resolved.unmapped_etabs_rows) == 1
    assert any("Unmapped ETABS demand rows: 1" in item.message for item in resolved.diagnostics)


def test_missing_catalog_reference_can_be_retained_as_diagnostic_when_configured(
    project_root: Path,
) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "members.xlsx"),
        "Members",
        "case_id",
        "EX_BEAM_001",
        section_id="UNKNOWN",
    )
    _edit_yaml(
        project_root,
        lambda document: document["quality_assurance"].__setitem__(  # type: ignore[index]
            "fail_on_missing_catalog_reference", False
        ),
    )
    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )
    assert resolved.active_resolved_members == ()
    assert any(item.code == "MISSING_SECTION_REFERENCE" for item in resolved.diagnostics)


def test_catalog_verification_runs_once_per_unique_active_section(
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _activate_example(project_root)
    members_path = _workbook(project_root, "members.xlsx")

    def add_second_member_and_mapping(workbook: object) -> None:
        members = workbook["Members"]  # type: ignore[index]
        member_headers = _headers(members)
        member_values = [cell.value for cell in members[2]]
        member_values[member_headers["case_id"] - 1] = "EX_BEAM_CLONE"
        member_values[member_headers["label"] - 1] = "Second active member"
        member_values[member_headers["active"] - 1] = True
        members.append(member_values)

        mappings = workbook["ETABS_Mapping"]  # type: ignore[index]
        mapping_headers = _headers(mappings)
        mappings.cell(2, mapping_headers["etabs_story"]).value = "FIRST ONLY"
        mappings.cell(2, mapping_headers["etabs_beam"]).value = "FIRST ONLY"
        mapping_values = [cell.value for cell in mappings[2]]
        mapping_values[mapping_headers["case_id"] - 1] = "EX_BEAM_CLONE"
        mapping_values[mapping_headers["etabs_unique_name"] - 1] = "2263"
        mapping_values[mapping_headers["etabs_story"] - 1] = "SECOND ONLY"
        mapping_values[mapping_headers["etabs_beam"] - 1] = "SECOND ONLY"
        mapping_values[mapping_headers["mapping_enabled"] - 1] = True
        mappings.append(mapping_values)

    _modify_workbook(members_path, add_second_member_and_mapping)

    def split_etabs(workbook: object) -> None:
        worksheet = workbook["Element Forces - Beams"]  # type: ignore[index]
        headers = _headers(worksheet, row=2)
        for row in range(16, 28):
            worksheet.cell(row, headers["Unique Name"]).value = "2263"

    _modify_workbook(_workbook(project_root, "ETABS_results.xlsx"), split_etabs)

    resolver_module = import_module("cfs_design.workflows.project.resolver")
    original = resolver_module.verify_catalog_properties
    calls = 0

    def counted(*args: object, **kwargs: object):
        nonlocal calls
        calls += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(resolver_module, "verify_catalog_properties", counted)
    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )
    assert calls == 1
    assert len(resolved.section_verification_results) == 1
    assert len(resolved.active_resolved_members) == 2


def test_required_missing_catalog_property_is_fatal(project_root: Path) -> None:
    _activate_example(project_root)

    def require_cw(document: dict[str, object]) -> None:
        verification = document["catalog_verification"]  # type: ignore[index]
        verification["required_properties"].append("Cw")
        verification["extended_properties"].remove("Cw")

    _edit_yaml(project_root, require_cw)
    with pytest.raises(ConfigurationError, match="missing required catalog values.*Cw"):
        resolve_project(_project_yaml(project_root), repository_root=project_root)


@pytest.mark.parametrize("action", ("warning", "error"))
def test_catalog_verification_fail_obeys_action(
    project_root: Path,
    action: str,
) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "sections_catalog.xlsx"),
        "Properties",
        "section_id",
        "EX_SEC_C200_70_20_2",
        A_mm2=5000.0,
    )
    _edit_yaml(
        project_root,
        lambda document: document["catalog_verification"].__setitem__(  # type: ignore[index]
            "action_on_fail", action
        ),
    )
    if action == "error":
        with pytest.raises(ConfigurationError, match="failed catalog verification"):
            resolve_project(_project_yaml(project_root), repository_root=project_root)
    else:
        resolved = resolve_project(
            _project_yaml(project_root), repository_root=project_root
        )
        assert len(resolved.active_resolved_members) == 1
        assert any(item.code == "CATALOG_PROPERTY_FAIL" for item in resolved.diagnostics)
        mechanics = resolved.get_section_mechanics("EX_SEC_C200_70_20_2")
        assert not mechanics.design_use_permitted
        assert mechanics.verification is resolved.section_verification_results[0]
        with pytest.raises(ConfigurationError, match="blocked from design use"):
            resolved.require_design_mechanics("EX_SEC_C200_70_20_2")


def test_disabled_catalog_verification_preserves_computed_set_but_blocks_design(
    project_root: Path,
) -> None:
    _activate_example(project_root)
    _edit_yaml(
        project_root,
        lambda document: document["catalog_verification"].__setitem__(  # type: ignore[index]
            "enabled", False
        ),
    )

    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )
    mechanics = resolved.get_section_mechanics("EX_SEC_C200_70_20_2")

    assert resolved.section_verification_results == ()
    assert mechanics.verification is None
    assert not mechanics.design_use_permitted
    assert isinstance(mechanics.gross, ComputedSectionProperties)
    assert isinstance(mechanics.advanced, AdvancedSectionProperties)
    with pytest.raises(ConfigurationError, match="blocked from design use"):
        resolved.require_design_mechanics("EX_SEC_C200_70_20_2")
    with pytest.raises(
        ValidationError,
        match="cannot be permitted without catalog verification",
    ):
        replace(mechanics, design_use_permitted=True)


@pytest.mark.parametrize("action", ("warning", "error"))
def test_unsupported_verification_geometry_obeys_action(
    project_root: Path,
    action: str,
) -> None:
    _activate_example(project_root)
    _set_by_id(
        _workbook(project_root, "sections_catalog.xlsx"),
        "Geometry",
        "geometry_id",
        "EX_GEO_C200",
        geometry_convention="OUT_TO_OUT",
    )
    _edit_yaml(
        project_root,
        lambda document: document["catalog_verification"].__setitem__(  # type: ignore[index]
            "action_on_fail", action
        ),
    )
    if action == "error":
        with pytest.raises(ConfigurationError, match="could not be mechanically verified"):
            resolve_project(_project_yaml(project_root), repository_root=project_root)
    else:
        resolved = resolve_project(
            _project_yaml(project_root), repository_root=project_root
        )
        assert len(resolved.active_resolved_members) == 1
        assert resolved.section_verification_results == ()
        diagnostic = next(
            item
            for item in resolved.diagnostics
            if item.code == "CATALOG_VERIFICATION_UNSUPPORTED"
        )
        assert diagnostic.severity is DiagnosticSeverity.WARNING
        assert "OUT_TO_OUT" in diagnostic.message
