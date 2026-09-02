"""M8A.2 coherent design-input and execution-purpose boundary tests."""

from collections.abc import Callable
from dataclasses import fields, replace
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from cfs_design.catalogs.schemas import MATERIAL_QUALIFICATION_COLUMNS
from cfs_design.core.exceptions import ConfigurationError
from cfs_design.design import MemberDesignInput
from cfs_design.domain import DesignMethod
from cfs_design.normative import (
    DesignAction,
    DesignExecutionPurpose,
    SoftwareSupportStatus,
)
from cfs_design.results import ApplicabilityStatus
from cfs_design.workflows.project import (
    resolve_member_design_input,
    resolve_project,
)


def _project_yaml(root: Path) -> Path:
    return root / "projects" / "PRJ_001" / "project.yaml"


def _workbook(root: Path, name: str) -> Path:
    if name in {"materials_catalog.xlsx", "sections_catalog.xlsx"}:
        return root / "data" / "catalogs" / name
    return root / "projects" / "PRJ_001" / name


def _modify(path: Path, operation: Callable[[object], None]) -> None:
    workbook = load_workbook(path)
    operation(workbook)
    workbook.save(path)
    workbook.close()


def _headers(worksheet: object) -> dict[object, int]:
    return {cell.value: cell.column for cell in worksheet[1]}  # type: ignore[index]


def _set_by_id(
    path: Path,
    sheet: str,
    id_header: str,
    identifier: str,
    **updates: object,
) -> None:
    def operation(workbook: object) -> None:
        worksheet = workbook[sheet]  # type: ignore[index]
        headers = _headers(worksheet)
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, headers[id_header]).value == identifier:
                for header, value in updates.items():
                    worksheet.cell(row, headers[header]).value = value
                return
        raise AssertionError(f"No {identifier} in {sheet}")

    _modify(path, operation)


def _configure_scope(root: Path) -> None:
    path = _project_yaml(root)
    document = yaml.safe_load(path.read_text(encoding="utf-8"))
    evidence = document["aisi_scope_evidence"]
    evidence["governing_country"]["country"] = "UNITED_STATES"
    evidence["structure_application"]["application"] = "BUILDING"
    evidence["cold_formed_to_shape"]["state"] = "TRUE"
    evidence["structural_load_carrying_use"]["state"] = "TRUE"
    path.write_text(yaml.safe_dump(document, sort_keys=False), encoding="utf-8")


def _upgrade_materials_with_qualification(root: Path) -> None:
    path = _workbook(root, "materials_catalog.xlsx")

    def operation(workbook: object) -> None:
        metadata = workbook["Metadata"]  # type: ignore[index]
        metadata_headers = _headers(metadata)
        for row in range(2, metadata.max_row + 1):
            if metadata.cell(row, metadata_headers["Field"]).value == "schema_version":
                metadata.cell(row, metadata_headers["Value"]).value = "0.2.0"
                break
        if "AISI_Material_Qualification" in workbook.sheetnames:  # type: ignore[attr-defined]
            qualification = workbook["AISI_Material_Qualification"]  # type: ignore[index]
            assert tuple(cell.value for cell in qualification[1]) == (
                MATERIAL_QUALIFICATION_COLUMNS
            )
        else:
            qualification = workbook.create_sheet(  # type: ignore[attr-defined]
                "AISI_Material_Qualification", 2
            )
            qualification.append(MATERIAL_QUALIFICATION_COLUMNS)
        values = {
            "material_id": "EX_MAT_G50",
            "standard_id": "ANSI_SDI_AISI_S100",
            "standard_edition": 2024,
            "qualification_route": "A3_1",
            "qualification_state": "QUALIFIED",
            "product_form": "SHEET",
            "steel_classification": "CARBON",
            "elongation_group": "A3_1_1_GE_10",
            "minimum_elongation_percent": 10.0,
            "elongation_gauge_length_mm": 50.0,
            "elongation_test_standard": "ASTM_A370",
            "mandatory_mechanical_properties_state": "SATISFIED",
            "test_reports_required_state": "SATISFIED",
            "chemical_mechanical_conformance_state": "NOT_APPLICABLE",
            "properties_determined_per_reference_state": "NOT_APPLICABLE",
            "coating_requirements_state": "NOT_APPLICABLE",
            "welding_requirements_state": "NOT_APPLICABLE",
            "production_identification_state": "NOT_APPLICABLE",
            "master_coil_10_percent_overstrength_state": "NOT_APPLICABLE",
            "local_elongation_percent": None,
            "uniform_elongation_percent": None,
            "ductility_test_standard": None,
            "source_id": "EX_SRC_MAT",
            "basis": "Synthetic verified qualification fixture.",
            "notes": "SYNTHETIC_TEST_DATA",
        }
        qualification.append(
            [values[column] for column in MATERIAL_QUALIFICATION_COLUMNS]
        )

    _modify(path, operation)
    _set_by_id(path, "Materials", "material_id", "EX_MAT_G50", active=True)


def _add_standard_dimensions(root: Path) -> None:
    path = _workbook(root, "sections_catalog.xlsx")

    def operation(workbook: object) -> None:
        worksheet = workbook["AISI_Dimensions"]  # type: ignore[index]
        values = {
            "geometry_id": "EX_GEO_C200",
            "standard_id": "ANSI_SDI_AISI_S100",
            "standard_edition": 2024,
            "web_flat_width_mm": 196.0,
            "flange_1_flat_width_mm": 66.0,
            "flange_2_flat_width_mm": 66.0,
            "web_out_to_out_depth_mm": 204.0,
            "flange_1_out_to_out_width_mm": 72.0,
            "flange_2_out_to_out_width_mm": 72.0,
            "lip_1_flat_width_mm": 16.0,
            "lip_2_flat_width_mm": 16.0,
            "lip_1_out_to_out_width_mm": 20.0,
            "lip_2_out_to_out_width_mm": 20.0,
            "lip_1_overall_depth_mm": 20.0,
            "lip_2_overall_depth_mm": 20.0,
            "source_id": "EX_SRC_SEC",
            "notes": "SYNTHETIC_TEST_DATA",
        }
        worksheet.append([values.get(cell.value) for cell in worksheet[1]])

    _modify(path, operation)
    _set_by_id(
        path,
        "Sections",
        "section_id",
        "EX_SEC_C200_70_20_2",
        active=True,
    )
    # Deliberately differ from M3 while staying inside the 1% QA tolerance.
    _set_by_id(
        path,
        "Properties",
        "section_id",
        "EX_SEC_C200_70_20_2",
        A_mm2=765.0,
    )


def _activate_column(root: Path) -> None:
    path = _workbook(root, "members.xlsx")
    _set_by_id(
        path,
        "Members",
        "case_id",
        "EX_BEAM_001",
        member_type="COLUMN",
        active=True,
        distortional_unbraced_length_mm=1800.0,
        distortional_restraint_source="Synthetic restraint schedule R-01.",
    )
    _set_by_id(
        path,
        "ETABS_Mapping",
        "case_id",
        "EX_BEAM_001",
        mapping_enabled=True,
    )


def _fully_resolved_project(root: Path):
    _configure_scope(root)
    _upgrade_materials_with_qualification(root)
    _add_standard_dimensions(root)
    _activate_column(root)
    return resolve_project(_project_yaml(root), repository_root=root)


def _resolve(project, purpose: DesignExecutionPurpose) -> MemberDesignInput:
    return resolve_member_design_input(
        project,
        "EX_BEAM_001",
        DesignMethod.EWM,
        DesignAction.AXIAL_COMPRESSION,
        purpose,
    )


def test_capacity_without_member_demands_can_be_executable(project_root: Path) -> None:
    resolved = _fully_resolved_project(project_root)
    original = resolved.active_resolved_members[0]
    no_demands = replace(original, demands=None, source_demands=None)
    capacity_project = replace(
        resolved,
        active_resolved_members=(no_demands,),
    )

    design_input = _resolve(capacity_project, DesignExecutionPurpose.CAPACITY)

    assert design_input.executable
    assert design_input.eligibility.normative.status is ApplicabilityStatus.APPLICABLE
    assert design_input.eligibility.software.status is SoftwareSupportStatus.SUPPORTED
    assert design_input.resolved_member.section_demands is None


def test_design_input_uses_one_coherent_m3_set_not_catalog_properties(
    project_root: Path,
) -> None:
    resolved = _fully_resolved_project(project_root)

    design_input = _resolve(resolved, DesignExecutionPurpose.CAPACITY)
    mechanics = resolved.require_design_mechanics("EX_SEC_C200_70_20_2")

    assert design_input.section_mechanics is mechanics
    assert design_input.section_mechanics.gross.a_mm2 == pytest.approx(760.0)
    assert design_input.resolved_member.section.properties.a_mm2 == 765.0
    assert design_input.section_mechanics.gross is not (
        design_input.resolved_member.section.properties
    )


def test_demand_check_without_m4_m5_pair_is_invalid(project_root: Path) -> None:
    resolved = _fully_resolved_project(project_root)
    no_demands = replace(
        resolved.active_resolved_members[0],
        demands=None,
        source_demands=None,
    )
    demand_project = replace(resolved, active_resolved_members=(no_demands,))

    design_input = _resolve(demand_project, DesignExecutionPurpose.DEMAND_CHECK)

    assert not design_input.executable
    assert design_input.eligibility.software.status is SoftwareSupportStatus.INVALID_INPUT


def test_demand_check_with_preserved_m4_m5_pair_can_be_executable(
    project_root: Path,
) -> None:
    resolved = _fully_resolved_project(project_root)

    design_input = _resolve(resolved, DesignExecutionPurpose.DEMAND_CHECK)

    assert design_input.executable
    assert design_input.resolved_member.source_demands is not None
    assert design_input.resolved_member.section_demands is not None


def test_failed_qa_gate_blocks_design_input(project_root: Path) -> None:
    _configure_scope(project_root)
    _upgrade_materials_with_qualification(project_root)
    _add_standard_dimensions(project_root)
    _activate_column(project_root)
    _set_by_id(
        _workbook(project_root, "sections_catalog.xlsx"),
        "Properties",
        "section_id",
        "EX_SEC_C200_70_20_2",
        A_mm2=5000.0,
    )
    resolved = resolve_project(
        _project_yaml(project_root), repository_root=project_root
    )

    with pytest.raises(ConfigurationError, match="blocked from design use"):
        _resolve(resolved, DesignExecutionPurpose.CAPACITY)


def test_member_design_input_has_no_resistance_or_utilization_fields() -> None:
    field_names = {item.name.lower() for item in fields(MemberDesignInput)}
    assert field_names.isdisjoint(
        {
            "fe",
            "fn",
            "pn",
            "mn",
            "phi",
            "phipn",
            "effective_width",
            "effective_area",
            "utilization",
        }
    )
